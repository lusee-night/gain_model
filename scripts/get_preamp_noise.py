#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
thermal_preamp_noise_model.py

Centered quadratic temperature model + log-frequency splines.

Model
-----
From Thermal test (anchor freqs f_j):
  N_th(T; f_j) = A_j + B_j*(T - T0) + C_j*(T - T0)^2

where T0 is the baseline temp (room temp) used to anchor scaling, default T0=25°C.

From Noise sheet (dense spectrum at ~T0):
  N_base(f) = spline over log10(f) through dense noise data

Temperature scaling factor at anchors:
  S(T, f_j) = N_th(T; f_j) / N_th(T0; f_j)

Interpolate S across log-frequency using only anchor freqs:
  S(T, f) = spline over log10(f) through (f_j, S(T, f_j))

Final model:
  N(T, f) = N_base(f) * S(T, f)

Exports
-------
Always:
  - prints N(Tq, fq) + scale factor and baseline at fq

Optional:
  --out <csv> writes a 1-row CSV

Plots (unless --no_plots), saved to:
  ~/gain_model/outputs/preamp_noise/plots/FMPRE<x>/
where <x> extracted from input filename like FMPRE1.xlsx

Plots include:
  - thermal fits per anchor frequency (data vs centered quadratic)
  - scale factor vs frequency at Tq
  - baseline vs scaled spectrum at Tq
  - baseline spectrum vs thermal model at T0 (anchors)
"""

import argparse
import csv
import os
import re
from typing import List, Tuple

import numpy as np
import openpyxl
import matplotlib
matplotlib.use("Agg")  # headless safe
import matplotlib.pyplot as plt
from scipy.interpolate import InterpolatedUnivariateSpline


# -----------------------------
# Configuration
# -----------------------------
T0_C = 25.0  # baseline temperature (room temp) for scaling and for centered polynomial
BASE_PLOTS_ROOT = os.path.expanduser("~/gain_model/outputs/preamp_noise/plots")


# -----------------------------
# Helpers: parsing & formatting
# -----------------------------
FREQ_RE = re.compile(r"^\s*([0-9]*\.?[0-9]+)\s*([kKmMgG]?)\s*[hH][zZ]\s*$")


def parse_frequency_to_hz(s: str) -> float:
    """Accepts '10MHz', '500kHz', '1e7', '0.5MHz' -> Hz float."""
    ss = str(s).strip()
    try:
        return float(ss)
    except ValueError:
        pass

    m = FREQ_RE.match(ss)
    if not m:
        raise ValueError(f"Could not parse frequency: '{s}' (try '10MHz' or '1e7').")

    val = float(m.group(1))
    prefix = (m.group(2) or "").lower()
    mult = {"": 1.0, "k": 1e3, "m": 1e6, "g": 1e9}[prefix]
    return val * mult


def parse_header_freq_to_hz(h: str) -> float:
    """Parses headers like '0.5MHz', '1MHz', '50MHz' -> Hz float."""
    hs = str(h).strip()
    m = re.match(r"^\s*([0-9]*\.?[0-9]+)\s*([kKmMgG]?)\s*[mM]?[hH][zZ]\s*$", hs)
    if not m:
        raise ValueError(f"Could not parse frequency header: '{h}'")
    val = float(m.group(1))
    prefix = (m.group(2) or "").lower()
    mult = {"": 1.0, "k": 1e3, "m": 1e6, "g": 1e9}[prefix]
    return val * mult


def format_freq_label(f_hz: float) -> str:
    if f_hz >= 1e9:
        return f"{f_hz/1e9:g}GHz"
    if f_hz >= 1e6:
        return f"{f_hz/1e6:g}MHz"
    if f_hz >= 1e3:
        return f"{f_hz/1e3:g}kHz"
    return f"{f_hz:g}Hz"


def safe_filename(s: str) -> str:
    return re.sub(r"[^A-Za-z0-9._-]+", "_", s)


def get_fmpre_plot_dir(xlsx_path: str) -> str:
    """~/gain_model/outputs/preamp_noise/plots/FMPRE<x> based on filename like FMPRE1.xlsx."""
    base = os.path.basename(xlsx_path)
    m = re.search(r"(FMPRE\d+)", base, re.IGNORECASE)
    tag = m.group(1).upper() if m else "FMPRE_UNKNOWN"
    return os.path.join(BASE_PLOTS_ROOT, tag)


# -----------------------------
# Excel extraction: Thermal test / Noise table
# -----------------------------
def find_noise_anchor(ws) -> Tuple[int, int]:
    """Find the cell containing exactly 'Noise' (case-insensitive)."""
    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            if isinstance(v, str) and v.strip().lower() == "noise":
                return cell.row, cell.column
    raise RuntimeError("Could not find a cell labeled 'Noise' in the sheet.")


def extract_thermal_noise_table(ws) -> Tuple[np.ndarray, np.ndarray, np.ndarray]:
    """
    Extract the Noise vs Temperature table from 'Thermal test' sheet.

    Expected format:
      Row after 'Noise': headers: Temperature | std | 0.5MHz | std | 1MHz | std | ...
      Then rows of numeric temperature and noise values.

    Returns:
      T: (nT,) temperatures in °C (sorted increasing)
      F: (nF,) anchor frequencies in Hz (sorted increasing)
      Z: (nT, nF) noise values in nV/√Hz
    """
    noise_row, _ = find_noise_anchor(ws)
    header_row = noise_row + 1

    headers = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(header_row, c).value
        if v is None:
            continue
        if isinstance(v, str):
            headers[c] = v.strip()

    temp_col = None
    freq_cols: List[Tuple[int, float]] = []

    for c, h in headers.items():
        if h.lower() == "temperature":
            temp_col = c
            continue
        if h.lower() == "std":
            continue
        try:
            f_hz = parse_header_freq_to_hz(h)
            freq_cols.append((c, f_hz))
        except Exception:
            pass

    if temp_col is None:
        raise RuntimeError("Could not find 'Temperature' column in the thermal Noise table.")
    if not freq_cols:
        raise RuntimeError("Could not find any frequency columns in the thermal Noise table.")

    freq_cols.sort(key=lambda x: x[1])
    F = np.array([f for _, f in freq_cols], dtype=float)

    T_list: List[float] = []
    Z_rows: List[List[float]] = []

    r = header_row + 1
    while r <= ws.max_row:
        t = ws.cell(r, temp_col).value
        if t is None or (isinstance(t, str) and t.strip() == ""):
            break
        if not isinstance(t, (int, float)):
            break

        row_vals = []
        ok = True
        for c, _f in freq_cols:
            v = ws.cell(r, c).value
            if v is None:
                ok = False
                break
            row_vals.append(float(v))

        if ok:
            T_list.append(float(t))
            Z_rows.append(row_vals)
        r += 1

    if len(T_list) < 3:
        raise RuntimeError("Not enough thermal rows to fit temperature model (need >= 3).")

    T = np.array(T_list, dtype=float)
    Z = np.array(Z_rows, dtype=float)

    order = np.argsort(T)
    T = T[order]
    Z = Z[order, :]

    return T, F, Z


# -----------------------------
# Excel extraction: Noise sheet (dense spectrum)
# -----------------------------
def extract_dense_noise_sheet(ws) -> Tuple[np.ndarray, np.ndarray]:
    """
    Extract dense spectrum from 'Noise' sheet.

    Finds columns by header containing:
      - 'freq' for frequency (Hz)
      - 'noise' for noise (nV/√Hz)
    """
    header_row = None
    freq_col = None
    noise_col = None

    max_scan_rows = min(ws.max_row, 50)

    for r in range(1, max_scan_rows + 1):
        row_vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        norm = [v.strip().lower() if isinstance(v, str) else None for v in row_vals]

        cand_freq = None
        cand_noise = None
        for c, v in enumerate(norm, start=1):
            if v is None:
                continue
            if cand_freq is None and "freq" in v:
                cand_freq = c
            if cand_noise is None and "noise" in v:
                cand_noise = c

        if cand_freq is not None and cand_noise is not None:
            header_row = r
            freq_col = cand_freq
            noise_col = cand_noise
            break

    if header_row is None or freq_col is None or noise_col is None:
        raise RuntimeError("Could not locate Frequency/Noise columns in 'Noise' sheet by header search.")

    F_list = []
    N_list = []

    r = header_row + 1
    while r <= ws.max_row:
        f = ws.cell(r, freq_col).value
        n = ws.cell(r, noise_col).value
        if f is None or n is None:
            break
        if not isinstance(f, (int, float)) or not isinstance(n, (int, float)):
            break
        F_list.append(float(f))
        N_list.append(float(n))
        r += 1

    if len(F_list) < 5:
        raise RuntimeError("Dense 'Noise' sheet did not yield enough rows (need >= 5).")

    F_dense = np.array(F_list, dtype=float)
    N_dense = np.array(N_list, dtype=float)

    order = np.argsort(F_dense)
    F_dense = F_dense[order]
    N_dense = N_dense[order]

    return F_dense, N_dense


# -----------------------------
# Centered quadratic temperature model
# -----------------------------
def eval_centered_quadratic(T: np.ndarray, A: float, B: float, C: float, T0: float) -> np.ndarray:
    """N(T) = A + B*(T-T0) + C*(T-T0)^2"""
    u = (T - T0)
    return A + B * u + C * (u ** 2)


def fit_centered_quadratic_per_frequency(T: np.ndarray, F_anchor: np.ndarray, Z: np.ndarray, T0: float) -> Tuple[np.ndarray, np.ndarray, np.ndarray]:
    """
    For each anchor frequency column y(T), fit:
      y = A + B*(T-T0) + C*(T-T0)^2
    via linear least squares.

    Returns arrays A_arr, B_arr, C_arr each length nF.
    """
    u = T - T0
    X = np.vstack([np.ones_like(u), u, u**2]).T  # (nT, 3)

    A_arr = np.zeros(len(F_anchor), dtype=float)
    B_arr = np.zeros(len(F_anchor), dtype=float)
    C_arr = np.zeros(len(F_anchor), dtype=float)

    for j in range(len(F_anchor)):
        y = Z[:, j].astype(float)
        coeffs, *_ = np.linalg.lstsq(X, y, rcond=None)  # [A,B,C]
        A_arr[j], B_arr[j], C_arr[j] = coeffs

    return A_arr, B_arr, C_arr

def rmse_and_nrmse(y: np.ndarray, yhat: np.ndarray) -> Tuple[float, float]:
    """Return (RMSE, NRMSE_mean). NRMSE_mean = RMSE / mean(y)."""
    y = np.asarray(y, dtype=float)
    yhat = np.asarray(yhat, dtype=float)
    rmse = float(np.sqrt(np.mean((y - yhat) ** 2)))
    denom = float(np.mean(np.abs(y)))
    nrmse = float(rmse / denom) if denom > 0 else float("nan")
    return rmse, nrmse


# -----------------------------
# Plotting
# -----------------------------
def save_temp_fit_plots(T, F_anchor, Z, A_arr, B_arr, C_arr, plot_dir, prefix=""):
    os.makedirs(plot_dir, exist_ok=True)
    T_dense = np.linspace(float(np.min(T)), float(np.max(T)), 400)

    for j, f_hz in enumerate(F_anchor):
        y = Z[:, j]

        # Model predictions at the actual sample temperatures (for metrics)
        yhat = eval_centered_quadratic(T, float(A_arr[j]), float(B_arr[j]), float(C_arr[j]), T0_C)

        # Smooth curve for plotting
        y_fit = eval_centered_quadratic(T_dense, float(A_arr[j]), float(B_arr[j]), float(C_arr[j]), T0_C)

        rmse, nrmse = rmse_and_nrmse(y, yhat)

        fig = plt.figure()
        plt.plot(T, y, marker="o", linestyle="")
        plt.plot(T_dense, y_fit)

        plt.xlabel("Temperature (°C)")
        plt.ylabel("Noise (nV/√Hz)")
        plt.title(f"Thermal fit at {format_freq_label(f_hz)}  |  centered quadratic about T0={T0_C:g}°C")

        # Add annotation box (top-left in axes coordinates)
        txt = f"RMSE = {rmse:.4g} nV/√Hz\nNRMSE = {100*nrmse:.2f}%"
        plt.gca().text(
            0.02, 0.98, txt,
            transform=plt.gca().transAxes,
            va="top", ha="left",
            bbox=dict(boxstyle="round", facecolor="white", alpha=0.8, edgecolor="0.6"),
            fontsize=9
        )

        plt.tight_layout()
        outname = f"{prefix}temp_fit_{safe_filename(format_freq_label(f_hz))}.png"
        plt.savefig(os.path.join(plot_dir, outname), dpi=200)
        plt.close(fig)



def save_scale_factor_plot(Tq, F_anchor, scale_at_anchor, scale_spline, f_query, S_query, plot_dir, prefix=""):
    os.makedirs(plot_dir, exist_ok=True)

    logF = np.log10(F_anchor)
    logF_dense = np.linspace(float(np.min(logF)), float(np.max(logF)), 600)
    F_dense = 10 ** logF_dense
    S_dense = scale_spline(logF_dense)

    fig = plt.figure()
    plt.plot(F_anchor, scale_at_anchor, marker="o", linestyle="")
    plt.plot(F_dense, S_dense)
    plt.plot([f_query], [S_query], marker="x", linestyle="")
    plt.xscale("log")
    plt.xlabel("Frequency (Hz)")
    plt.ylabel(f"Scale factor S(T,f) relative to {T0_C:g}°C")
    plt.title(f"Scale factor vs frequency at T={Tq:g}°C (log10(f) spline)")
    plt.tight_layout()

    outname = f"{prefix}scale_factor_at_T_{safe_filename(str(Tq))}C.png"
    plt.savefig(os.path.join(plot_dir, outname), dpi=200)
    plt.close(fig)


def save_spectrum_plot(Tq, F_dense, N_base_dense, scaled_dense, F_anchor, N_th_Tq, f_query, noise_q, plot_dir, prefix=""):
    os.makedirs(plot_dir, exist_ok=True)

    fig = plt.figure()
    plt.plot(F_dense, N_base_dense, label=f"Baseline (Noise sheet, ~{T0_C:g}°C)")
    plt.plot(F_dense, scaled_dense, label=f"Scaled to T={Tq:g}°C")
    plt.plot(F_anchor, N_th_Tq, marker="o", linestyle="", label="Thermal anchors at T (model)")
    plt.plot([f_query], [noise_q], marker="x", linestyle="", label="Query")

    plt.xscale("log")
    plt.xlabel("Frequency (Hz)")
    plt.ylabel("Noise (nV/√Hz)")
    plt.title(f"Baseline vs scaled spectrum at T={Tq:g}°C")
    plt.legend()
    plt.tight_layout()

    outname = f"{prefix}spectrum_at_T_{safe_filename(str(Tq))}C.png"
    plt.savefig(os.path.join(plot_dir, outname), dpi=200)
    plt.close(fig)


def save_baseline_vs_T0_plot(F_dense, N_base_dense, F_anchor, N_th_T0, plot_dir, prefix=""):
    """Extra diagnostic: baseline dense spectrum vs thermal-model prediction at T0 (anchors)."""
    os.makedirs(plot_dir, exist_ok=True)

    fig = plt.figure()
    plt.plot(F_dense, N_base_dense, label=f"Baseline (Noise sheet, ~{T0_C:g}°C)")
    plt.plot(F_anchor, N_th_T0, marker="o", linestyle="", label=f"Thermal model at T0={T0_C:g}°C (anchors)")

    plt.xscale("log")
    plt.xlabel("Frequency (Hz)")
    plt.ylabel("Noise (nV/√Hz)")
    plt.title(f"Baseline spectrum vs thermal model at T0 = {T0_C:g}°C")
    plt.legend()
    plt.tight_layout()
    outname = f"{prefix}baseline_vs_T0.png"
    plt.savefig(os.path.join(plot_dir, outname), dpi=200)
    plt.close(fig)

# -----------------------------
# Main
# -----------------------------
def main():
    ap = argparse.ArgumentParser(description="Preamp noise model: centered quadratic in T, baseline shape in f, log-f splines.")
    ap.add_argument("--xlsx", required=True, help="Input Excel file (e.g., FMPRE1.xlsx)")
    ap.add_argument("--thermal_sheet", default="Thermal test", help="Sheet containing thermal Noise table")
    ap.add_argument("--noise_sheet", default="Noise", help="Sheet containing dense noise spectrum")
    ap.add_argument("--temperature", type=float, required=True, help="Query temperature in °C")
    ap.add_argument("--frequency", required=True, help="Query frequency (e.g. 7.5MHz, 1e7)")
    ap.add_argument("--out", default=None, help="Optional output CSV path (query + result)")
    ap.add_argument(
        "--plot_dir",
        default=None,
        help="Override plot directory. Default: ~/gain_model/outputs/preamp_noise/plots/FMPRE<x>/"
    )
    ap.add_argument("--no_plots", action="store_true", help="Disable plot exports")
    ap.add_argument("--no_extrapolate", action="store_true",
                    help="Error out if query T or f is outside supported ranges.")
    args = ap.parse_args()

    Tq = float(args.temperature)
    f_query = parse_frequency_to_hz(args.frequency)

    plot_dir = args.plot_dir if args.plot_dir else get_fmpre_plot_dir(args.xlsx)

    wb = openpyxl.load_workbook(args.xlsx, data_only=True)

    if args.thermal_sheet not in wb.sheetnames:
        raise RuntimeError(f"Thermal sheet '{args.thermal_sheet}' not found. Available: {wb.sheetnames}")
    if args.noise_sheet not in wb.sheetnames:
        raise RuntimeError(f"Noise sheet '{args.noise_sheet}' not found. Available: {wb.sheetnames}")

    ws_th = wb[args.thermal_sheet]
    ws_ns = wb[args.noise_sheet]

    # ---- Read data ----
    T, F_anchor, Z = extract_thermal_noise_table(ws_th)
    F_dense, N_dense = extract_dense_noise_sheet(ws_ns)

    # Safety: ensure T0 is within thermal range
    if not (np.min(T) <= T0_C <= np.max(T)):
        raise ValueError(f"T0={T0_C}°C is outside thermal temperature range [{np.min(T)}, {np.max(T)}].")

    if args.no_extrapolate:
        if not (np.min(T) <= Tq <= np.max(T)):
            raise ValueError(f"T={Tq}°C outside measured thermal range [{np.min(T)}, {np.max(T)}].")
        if not (np.min(F_dense) <= f_query <= np.max(F_dense)):
            raise ValueError(f"f={f_query} Hz outside dense Noise sheet range [{np.min(F_dense)}, {np.max(F_dense)}].")

    # ---- Fit centered quadratic per anchor frequency ----
    A_arr, B_arr, C_arr = fit_centered_quadratic_per_frequency(T, F_anchor, Z, T0_C)

    # Evaluate thermal model at Tq and T0 for each anchor frequency
    N_th_Tq = eval_centered_quadratic(np.array([Tq]), 1.0, 0.0, 0.0, T0_C)  # dummy shape
    # compute vectorized (avoid loops)
    u_q = Tq - T0_C
    u_0 = 0.0  # T0 - T0

    N_th_Tq = A_arr + B_arr * u_q + C_arr * (u_q ** 2)
    N_th_T0 = A_arr + B_arr * u_0 + C_arr * (u_0 ** 2)  # = A_arr

    # Scale factors at anchor freqs
    scale_at_anchor = N_th_Tq / N_th_T0

    # ---- Scale factor spline over log10(f) ----
    logF_anchor = np.log10(F_anchor)
    k_scale = min(3, len(F_anchor) - 1)
    scale_spline = InterpolatedUnivariateSpline(logF_anchor, scale_at_anchor, k=k_scale)
    S_query = float(scale_spline(np.log10(f_query)))

    # ---- Baseline noise spline over log10(f) ----
    logF_dense = np.log10(F_dense)
    k_base = min(3, len(F_dense) - 1)
    base_spline = InterpolatedUnivariateSpline(logF_dense, N_dense, k=k_base)
    N_base_query = float(base_spline(np.log10(f_query)))

    noise_q = N_base_query * S_query

    print(f"Noise(T={Tq:.6g} °C, f={f_query:.6g} Hz) = {noise_q:.6g} nV/√Hz")
    print(f"  (T0={T0_C:g}°C baseline; S={S_query:.6g}; baseline_at_f={N_base_query:.6g})")

    # Optional CSV output
    if args.out:
        with open(args.out, "w", newline="") as fp:
            w = csv.writer(fp)
            w.writerow([
                "xlsx",
                "thermal_sheet",
                "noise_sheet",
                "temperature_C",
                "frequency_Hz",
                "noise_nV_per_sqrtHz",
                "T0_C",
                "scale_factor_S",
                "baseline_noise_at_f"
            ])
            w.writerow([
                args.xlsx,
                args.thermal_sheet,
                args.noise_sheet,
                Tq,
                f_query,
                noise_q,
                T0_C,
                S_query,
                N_base_query
            ])

    # ---- Plots ----
    if not args.no_plots:
        os.makedirs(plot_dir, exist_ok=True)
        base = os.path.splitext(os.path.basename(args.xlsx))[0]
        prefix = safe_filename(base) + "_"

        # 1) thermal fits per anchor frequency
        save_temp_fit_plots(T, F_anchor, Z, A_arr, B_arr, C_arr, plot_dir, prefix=prefix)

        # 2) scale factor vs frequency at Tq
        save_scale_factor_plot(Tq, F_anchor, scale_at_anchor, scale_spline, f_query, S_query, plot_dir, prefix=prefix)

        # 3) spectrum: baseline vs scaled at Tq (scaled on dense grid)
        S_dense = scale_spline(np.log10(F_dense))
        scaled_dense = N_dense * S_dense
        save_spectrum_plot(Tq, F_dense, N_dense, scaled_dense, F_anchor, N_th_Tq, f_query, noise_q, plot_dir, prefix=prefix)

        # 4) extra diagnostic: baseline vs thermal model at T0
        save_baseline_vs_T0_plot(F_dense, N_dense, F_anchor, N_th_T0, plot_dir, prefix=prefix)

        print(f"Saved diagnostic plots to: {plot_dir}")


if __name__ == "__main__":
    main()
