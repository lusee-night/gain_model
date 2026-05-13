#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Noise SVD on grouped LuSEE session HDF5 spectra, grouped by notch-defined mod-4 phase,
stacked across all science campaigns.

Original workflow
-----------------
- Reads session_science_###.h5 files recursively from:
      ~/gain_model/data/h5/Payload_CPTs/Payload_CPTs_Science_h5/
- Ignores empty or malformed files
- Assigns valid files to notch-defined phases and keeps only nmod4=1,2,3 for analysis
- For each retained mod-4 group and product:
    - loads spectra and rescales each product by 2**(actual_bitslice[p] - 31)
    - stacks all spectra rows across all campaigns/files in that group
    - removes non-finite rows
    - computes mean and median spectra
    - computes bad bins from the ORIGINAL stacked matrix exactly as before
    - runs SVD on:
        * fractional residuals for products 0-3
        * absolute residuals for products 4-15
    - saves mean plots, residual examples, eigenvalues, eigvecs, and PC scores

New workflow added here (products 0-3 only)
-------------------------------------------
For each row in the stacked matrix:
    - extract telemetry fields needed by the corrected-gain model
    - extract gain level (L/M/H)
    - compute corrected gain at every bin frequency using cubic spline interpolation
    - compute preamp noise at every bin frequency using PFPS_PA{x}_T
    - convert the original spectra to nV/sqrt(Hz) via:
          autos:  sqrt(original_spectrum / corrected_gain)
          cross:  sign(original_spectrum) * sqrt(abs(original_spectrum) / sqrt(gain_a * gain_b))

Then:
    - compare/subtract modeled preamp contribution against the converted nV/sqrt(Hz) spectra
    - compute and save a plot of:
          converted mean spectrum  vs  mean modeled preamp-noise spectrum
      using the ORIGINAL masked bins
    - subtract the modeled preamp spectrum row-by-row:
          X_preamp_subtracted = X_nV_per_sqrt_Hz - preamp_noise
    - zero out the ORIGINAL masked bins in X_preamp_subtracted
    - run a second SVD on this preamp-subtracted matrix
    - save analogous outputs with names prefixed by "preamp_subtracted_"

Notes
-----
- Bad-bin masking is determined ONLY from the original stacked matrix.
- For the preamp-subtracted branch, those same masked bins are simply set to 0.
- Products 4-15 are also converted to nV/sqrt(Hz) before the first SVD.
- Run with: PYTHONPATH=receive uv run ~/gain_model/scripts/noise_model.py
"""

import argparse
import csv
import glob
import os
import re
import h5py
from collections import defaultdict
from typing import Dict, List, Optional, Tuple

import numpy as np
import openpyxl
import pandas as pd

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from matplotlib.backends.backend_pdf import PdfPages

from scipy.interpolate import CubicSpline, InterpolatedUnivariateSpline

# NOTE: Run with: PYTHONPATH=receive uv run ~/gain_model/scripts/noise_model.py
try:
    from data import Spectra, TELEMETRY_FIELD_ALIASES
except ImportError:
    from data import Spectra
    TELEMETRY_FIELD_ALIASES = {}


_SESSION_NUM_RE = re.compile(r"session_science_(\d{3})\.h5$")
_CAMPAIGN_NUM_RE = re.compile(r"science_(\d+)$")

PRODUCT_NAMES = {
    0: "Ch0 Auto",
    1: "Ch1 Auto",
    2: "Ch2 Auto",
    3: "Ch3 Auto",
    4: "Ch0×Ch1 Re",
    5: "Ch0×Ch1 Im",
    6: "Ch0×Ch2 Re",
    7: "Ch0×Ch2 Im",
    8: "Ch0×Ch3 Re",
    9: "Ch0×Ch3 Im",
    10: "Ch1×Ch2 Re",
    11: "Ch1×Ch2 Im",
    12: "Ch1×Ch3 Re",
    13: "Ch1×Ch3 Im",
    14: "Ch2×Ch3 Re",
    15: "Ch2×Ch3 Im",
}
CROSS_PRODUCT_CHANNELS = {
    4: (0, 1),
    5: (0, 1),
    6: (0, 2),
    7: (0, 2),
    8: (0, 3),
    9: (0, 3),
    10: (1, 2),
    11: (1, 2),
    12: (1, 3),
    13: (1, 3),
    14: (2, 3),
    15: (2, 3),
}

CHANNEL_BIN_MHZ = 0.025
T0_C = 25.0
THRESHOLD_RATIO = 2.0
JACKKNIFE_MIN_ROWS = 30

# Defaults for gain-model products
DEFAULT_GAIN_MODEL_ROOT = os.path.expanduser("~/gain_model/outputs/gain_pca_model")
DEFAULT_GAIN_PREAMP_DIR = os.path.join(DEFAULT_GAIN_MODEL_ROOT, "corrected", "phase1", "pca")
DEFAULT_GAIN_ALPHAS_DIR = os.path.join(DEFAULT_GAIN_MODEL_ROOT, "corrected", "phase2", "alphas")

# Defaults for preamp-noise Excel files
DEFAULT_PREAMP_XLSX = {
    0: os.path.expanduser("~/gain_model/scripts/fmpre_noise/FMPRE6.xlsx"),
    1: os.path.expanduser("~/gain_model/scripts/fmpre_noise/FMPRE3.xlsx"),
    2: os.path.expanduser("~/gain_model/scripts/fmpre_noise/FMPRE4.xlsx"),
    3: os.path.expanduser("~/gain_model/scripts/fmpre_noise/FMPRE1.xlsx"),
}

REQUIRED_GAIN_TELEMETRY = [
    "THERM_FPGA",
    "SPE_ADC0_T",
    "SPE_ADC1_T",
    "SPE_1VAD8_V",
    "VMON_1V2D",
    "SPE_1VAD8_C",
]

REQUIRED_PREAMP_TEMP_FIELDS = [
    "PFPS_PA0_T",
    "PFPS_PA1_T",
    "PFPS_PA2_T",
    "PFPS_PA3_T",
]

ALL_REQUIRED_TELEMETRY = REQUIRED_GAIN_TELEMETRY + REQUIRED_PREAMP_TEMP_FIELDS


# ----------------------------
# Discovery / I/O helpers
# ----------------------------

def discover_session_files(session_dir: str, pattern: str) -> List[str]:
    session_dir = os.path.expanduser(session_dir)
    recursive_pattern = os.path.join(session_dir, "**", pattern)
    paths = sorted(glob.glob(recursive_pattern, recursive=True))
    return [p for p in paths if os.path.isfile(p)]


def ensure_dir(path: str) -> None:
    os.makedirs(path, exist_ok=True)


def write_csv_rows(path: str, header: List[str], rows: List[List[object]]) -> None:
    with open(path, "w", newline="") as f:
        w = csv.writer(f)
        w.writerow(header)
        for r in rows:
            w.writerow(r)


def extract_session_number(path: str) -> int:
    m = _SESSION_NUM_RE.search(os.path.basename(path))
    if not m:
        raise ValueError(f"Could not parse session number from filename: {path}")
    return int(m.group(1))


def extract_campaign_name(path: str, session_dir: str) -> str:
    session_dir = os.path.abspath(os.path.expanduser(session_dir))
    path = os.path.abspath(path)
    rel = os.path.relpath(path, session_dir)
    parts = rel.split(os.sep)
    return parts[0] if len(parts) > 1 else "unknown_campaign"


def extract_campaign_number(path: str, session_dir: str) -> int:
    campaign_name = extract_campaign_name(path, session_dir)
    m = _CAMPAIGN_NUM_RE.match(campaign_name)
    if not m:
        return -1
    return int(m.group(1))


def product_dir_name(p: int) -> str:
    raw = PRODUCT_NAMES.get(p, f"prod{p}")
    safe = (
        raw.replace("×", "x")
           .replace(" ", "_")
           .replace("/", "_")
           .replace("-", "_")
    )
    return f"prod_{p:03d}_{safe}"


def product_label(p: int) -> str:
    return PRODUCT_NAMES.get(p, f"Product {p}")

def short_title(group_name: str, product: int, label: str, gain_level: Optional[str] = None) -> str:
    m = re.search(r"(\d+)$", str(group_name))
    gtxt = f"n{m.group(1)}" if m else str(group_name)
    parts = [gtxt, f"P{product}"]
    if gain_level is not None and str(gain_level) != "":
        parts.append(str(gain_level))
    parts.append(label)
    return " • ".join(parts)


def short_stage_title(title_prefix: str, label: str) -> str:
    left = str(title_prefix).split(" — ", 1)[0].strip()
    return f"{left} • {label}"


def get_rowwise_bitslice_scale_from_h5(
    h5_path: str,
    expected_n_rows: int,
    n_prod: int,
) -> Tuple[np.ndarray, np.ndarray]:
    """
    Read actual_bitslice directly from each item_###/meta group in the HDF5 file,
    tile the per-product scale across that item's spectra rows, and concatenate
    the result in item order so it aligns with Spectra(path).data row order.

    Returns
    -------
    bitslice_rows : (expected_n_rows, n_prod)
        Per-row actual_bitslice values.
    scale_rows : (expected_n_rows, n_prod)
        Per-row multiplicative scale factors 2**(bitslice - 31).
    """
    bitslice_rows = []
    scale_rows = []

    with h5py.File(h5_path, "r") as f:
        item_names = sorted([k for k in f.keys() if k.startswith("item_")])

        for item in item_names:
            spectra_path = f"{item}/spectra/data"
            meta_path = f"{item}/meta"
            if spectra_path not in f or meta_path not in f:
                continue

            data_ds = f[spectra_path]
            if data_ds.ndim != 3:
                continue

            n_item_rows, n_item_prod, _n_bins = data_ds.shape
            if n_item_rows == 0:
                continue
            if n_item_prod != n_prod:
                raise ValueError(
                    f"Item {item} in {h5_path} has n_prod={n_item_prod}, expected {n_prod}"
                )

            meta = f[meta_path]
            if "actual_bitslice" in meta:
                bitslic = np.asarray(meta["actual_bitslice"], dtype=float).reshape(-1)
            elif "bitslice" in meta:
                bitslic = np.asarray(meta["bitslice"], dtype=float).reshape(-1)
            else:
                raise KeyError(f"Neither actual_bitslice nor bitslice found in {meta_path}")

            if bitslic.size != n_prod:
                raise ValueError(
                    f"Bitslice length {bitslic.size} in {meta_path}, expected {n_prod}"
                )
            if not np.all(np.isfinite(bitslic)):
                raise ValueError(f"Non-finite bitslice values in {meta_path}")

            scale = 2.0 ** (bitslic - 31.0)
            bitslice_rows.append(np.tile(bitslic.reshape(1, n_prod), (n_item_rows, 1)))
            scale_rows.append(np.tile(scale.reshape(1, n_prod), (n_item_rows, 1)))

    if not scale_rows:
        raise RuntimeError(f"No item_###/spectra/data groups found in {h5_path}")

    bitslice_rows = np.vstack(bitslice_rows)
    scale_rows = np.vstack(scale_rows)

    if bitslice_rows.shape != (expected_n_rows, n_prod):
        raise ValueError(
            f"Per-item bitslice rows shape {bitslice_rows.shape} does not match "
            f"Spectra data shape ({expected_n_rows}, {n_prod}) for {h5_path}"
        )

    return bitslice_rows, scale_rows


# ----------------------------
# Math helpers
# ----------------------------

def fractional_residuals(
    X: np.ndarray,
    mask_bins: Optional[np.ndarray] = None
) -> Tuple[np.ndarray, np.ndarray]:
    """
    X: (n_samples, n_bins)
    mu: mean over samples (n_bins,)
    R: (X - mu)/mu, with mu==0 treated as masked -> residuals forced to 0
    """
    mu = np.mean(X.astype(float), axis=0)

    if mask_bins is not None and mask_bins.size > 0:
        valid = (mask_bins >= 0) & (mask_bins < mu.size)
        mu = mu.copy()
        mu[mask_bins[valid]] = 0.0

    denom = mu.copy()
    zero_like = (denom == 0.0)
    denom_safe = denom.copy()
    denom_safe[zero_like] = 1.0

    R = (X.astype(float) - mu) / denom_safe
    R[:, zero_like] = 0.0
    return R, mu


def absolute_residuals(
    X: np.ndarray,
    mask_bins: Optional[np.ndarray] = None
) -> Tuple[np.ndarray, np.ndarray]:
    """
    X: (n_samples, n_bins)
    mu: mean over samples (n_bins,)
    R: X - mu, with masked bins forced to 0.
    """
    mu = np.mean(X.astype(float), axis=0)
    R = X.astype(float) - mu

    if mask_bins is not None and mask_bins.size > 0:
        valid = (mask_bins >= 0) & (mask_bins < mu.size)
        mu = mu.copy()
        R = R.copy()
        mu[mask_bins[valid]] = 0.0
        R[:, mask_bins[valid]] = 0.0

    return R, mu


def sanitize_matrix_rows(X: np.ndarray) -> Tuple[np.ndarray, np.ndarray]:
    keep_mask = np.all(np.isfinite(X), axis=1)
    X_clean = X[keep_mask]
    return X_clean, keep_mask


def svd_pca(R: np.ndarray) -> Tuple[np.ndarray, np.ndarray, np.ndarray]:
    if R.ndim != 2:
        raise ValueError(f"R must be 2D, got shape {R.shape}")
    if R.shape[0] == 0:
        raise ValueError("R has zero rows after cleaning")
    if not np.all(np.isfinite(R)):
        raise ValueError("R contains NaN or inf before SVD")

    U, S, Vt = np.linalg.svd(R, full_matrices=False)
    pcs = U * S
    eigvecs = Vt.T
    n = R.shape[0]
    eigenvals = (S ** 2) / (n - 1) if n > 1 else (S ** 2)
    return pcs, eigvecs, eigenvals


def moving_median_1d(y: np.ndarray, half_window: int) -> np.ndarray:
    """
    Simple 1D moving median.
    """
    y = np.asarray(y, dtype=float)
    n = y.size
    out = np.empty(n, dtype=float)

    for i in range(n):
        lo = max(0, i - half_window)
        hi = min(n, i + half_window + 1)
        out[i] = np.median(y[lo:hi])

    return out


def find_true_runs(mask: np.ndarray) -> List[Tuple[int, int]]:
    """Return inclusive (start, end) index pairs for contiguous True runs."""
    mask = np.asarray(mask, dtype=bool)
    if mask.size == 0:
        return []

    runs: List[Tuple[int, int]] = []
    in_run = False
    start = 0

    for i, val in enumerate(mask):
        if val and not in_run:
            start = i
            in_run = True
        elif not val and in_run:
            runs.append((start, i - 1))
            in_run = False

    if in_run:
        runs.append((start, mask.size - 1))

    return runs


def moving_average_1d(x: np.ndarray, window: int) -> np.ndarray:
    """Simple centered moving average with edge padding."""
    x = np.asarray(x, dtype=float)
    window = max(1, int(window))
    if window == 1:
        return x.copy()

    pad_left = window // 2
    pad_right = window - 1 - pad_left
    xpad = np.pad(x, (pad_left, pad_right), mode="edge")
    kernel = np.ones(window, dtype=float) / float(window)
    return np.convolve(xpad, kernel, mode="valid")


def detect_bad_bins_auto(
    X: np.ndarray,
    manual_mask_bins: Optional[np.ndarray] = None,
    start_bin: int = 150,
    smooth_window: int = 13,
    baseline_half_window: int = 61,
    absolute_stat_threshold: float = 1.0,
    iterative_remaining_threshold: float = 0.25,
    min_run_length: int = 2,
    mask_pad: int = 0,
    iterative_pad: int = 0,
) -> Tuple[np.ndarray, np.ndarray, np.ndarray, np.ndarray, str]:
    """
    Autocorrelation detector (0-3) using the same localized-excess statistic
    as the cross-correlation products so the plotted filter statistic is on the
    same footing across all products.
    """
    X = np.asarray(X, dtype=float)
    if X.ndim != 2:
        raise ValueError(f"X must be 2D, got shape {X.shape}")

    n_bins = X.shape[1]
    start_bin = max(0, int(start_bin))
    min_run_length = max(1, int(min_run_length))
    mask_pad = max(0, int(mask_pad))
    iterative_pad = max(0, int(iterative_pad))

    mean_spec = np.mean(X, axis=0)
    median_spec = np.median(X, axis=0)

    diff = np.abs(mean_spec - median_spec)
    diff_smooth = moving_average_1d(diff, smooth_window)
    baseline = moving_median_1d(diff_smooth, half_window=baseline_half_window)
    stat = np.maximum(diff_smooth - baseline, 0.0)

    raw_flag = np.zeros(n_bins, dtype=bool)
    if start_bin < n_bins:
        raw_flag[start_bin:] = stat[start_bin:] > float(absolute_stat_threshold)

    run_mask = np.zeros(n_bins, dtype=bool)
    for a, b in find_true_runs(raw_flag):
        if (b - a + 1) >= min_run_length:
            lo = max(0, a - mask_pad)
            hi = min(n_bins - 1, b + mask_pad)
            run_mask[lo:hi + 1] = True

    final_mask = run_mask.copy()

    if start_bin < n_bins:
        cleanup_threshold = float(iterative_remaining_threshold)
        while True:
            remaining_idx = np.where(~final_mask)[0]
            remaining_idx = remaining_idx[remaining_idx >= start_bin]
            if remaining_idx.size == 0:
                break

            rightmost_bin = int(remaining_idx.max())
            rightmost_val = float(stat[rightmost_bin])

            if rightmost_val <= cleanup_threshold:
                break

            lo = max(start_bin, rightmost_bin - iterative_pad)
            hi = min(n_bins - 1, rightmost_bin)
            final_mask[lo:hi + 1] = True

    if manual_mask_bins is not None and len(manual_mask_bins) > 0:
        manual_mask_bins = np.asarray(manual_mask_bins, dtype=int)
        valid = (manual_mask_bins >= 0) & (manual_mask_bins < n_bins)
        final_mask[manual_mask_bins[valid]] = True

    return np.where(final_mask)[0], mean_spec, median_spec, stat, (
        "Localized excess in smoothed |mean-median| above moving-median baseline "
        f"(absolute cutoff > {float(absolute_stat_threshold):.1f}; "
        f"right-edge trimming until highest remaining bin <= {float(iterative_remaining_threshold):.1f})"
    )


def detect_bad_bins_cross(
    X: np.ndarray,
    manual_mask_bins: Optional[np.ndarray] = None,
    start_bin: int = 150,
    smooth_window: int = 13,
    baseline_half_window: int = 61,
    absolute_stat_threshold: float = 1.0,
    iterative_remaining_threshold: float = 0.25,
    min_run_length: int = 2,
    mask_pad: int = 0,
    iterative_pad: int = 0,
) -> Tuple[np.ndarray, np.ndarray, np.ndarray, np.ndarray, str]:
    """
    Conservative detector for cross-correlation products (4-15).

    Build a localized-excess statistic in smoothed |mean - median| above a
    moving-median baseline. First mask runs where the statistic exceeds a fixed
    absolute threshold. Then trim only from the right edge until the highest
    remaining bin is at or below a second threshold.
    """
    X = np.asarray(X, dtype=float)
    if X.ndim != 2:
        raise ValueError(f"X must be 2D, got shape {X.shape}")

    n_bins = X.shape[1]
    start_bin = max(0, int(start_bin))
    min_run_length = max(1, int(min_run_length))
    mask_pad = max(0, int(mask_pad))
    iterative_pad = max(0, int(iterative_pad))

    mean_spec = np.mean(X, axis=0)
    median_spec = np.median(X, axis=0)

    diff = np.abs(mean_spec - median_spec)
    diff_smooth = moving_average_1d(diff, smooth_window)
    baseline = moving_median_1d(diff_smooth, half_window=baseline_half_window)
    stat = np.maximum(diff_smooth - baseline, 0.0)

    raw_flag = np.zeros(n_bins, dtype=bool)
    if start_bin < n_bins:
        raw_flag[start_bin:] = stat[start_bin:] > float(absolute_stat_threshold)

    run_mask = np.zeros(n_bins, dtype=bool)
    for a, b in find_true_runs(raw_flag):
        if (b - a + 1) >= min_run_length:
            lo = max(0, a - mask_pad)
            hi = min(n_bins - 1, b + mask_pad)
            run_mask[lo:hi + 1] = True

    final_mask = run_mask.copy()

    if start_bin < n_bins:
        cleanup_threshold = float(iterative_remaining_threshold)
        while True:
            remaining_idx = np.where(~final_mask)[0]
            remaining_idx = remaining_idx[remaining_idx >= start_bin]
            if remaining_idx.size == 0:
                break

            rightmost_bin = int(remaining_idx.max())
            rightmost_val = float(stat[rightmost_bin])

            if rightmost_val <= cleanup_threshold:
                break

            lo = max(start_bin, rightmost_bin - iterative_pad)
            hi = min(n_bins - 1, rightmost_bin)
            final_mask[lo:hi + 1] = True

    if manual_mask_bins is not None and len(manual_mask_bins) > 0:
        manual_mask_bins = np.asarray(manual_mask_bins, dtype=int)
        valid = (manual_mask_bins >= 0) & (manual_mask_bins < n_bins)
        final_mask[manual_mask_bins[valid]] = True

    return np.where(final_mask)[0], mean_spec, median_spec, stat, (
        "Localized excess in smoothed |mean-median| above moving-median baseline "
        f"(absolute cutoff > {float(absolute_stat_threshold):.1f}; "
        f"right-edge trimming until highest remaining bin <= {float(iterative_remaining_threshold):.1f})"
    )


# ----------------------------
# Plotting
# ----------------------------

def plot_mean_spectrum(
    mean_vec: np.ndarray,
    out_png: str,
    title: str,
    pdf: Optional[PdfPages] = None,
    ylabel: str = "Mean power",
    force_zero_bottom: bool = False,
) -> None:
    x = np.arange(mean_vec.size) * CHANNEL_BIN_MHZ
    y = mean_vec.astype(float).copy()
    y[y == 0.0] = np.nan

    fig = plt.figure(figsize=(9, 5.5))
    plt.plot(x, y)
    plt.xlabel("Frequency (MHz)")
    plt.ylabel(ylabel)
    plt.title(title)
    plt.grid(True, alpha=0.3)
    if force_zero_bottom:
        plt.ylim(bottom=0)
    plt.tight_layout()

    fig.savefig(out_png)
    if pdf is not None:
        pdf.savefig(fig)

    plt.close(fig)


def plot_mean_median_with_mask(
    mean_vec: np.ndarray,
    median_vec: np.ndarray,
    masked_bins: np.ndarray,
    out_png: str,
    title: str,
    ylabel: str = "Power",
    force_zero_bottom: bool = False,
) -> None:
    x = np.arange(mean_vec.size) * CHANNEL_BIN_MHZ
    y_mean = np.asarray(mean_vec, dtype=float).copy()
    y_median = np.asarray(median_vec, dtype=float).copy()

    plt.figure(figsize=(10, 5.5))
    plt.plot(x, y_mean, label="Mean spectrum", linewidth=1.3)
    plt.plot(x, y_median, label="Median spectrum", linewidth=1.3)

    if masked_bins.size > 0:
        mb = masked_bins[(masked_bins >= 0) & (masked_bins < mean_vec.size)]
        plt.scatter(x[mb], y_mean[mb], s=18, label="Masked bins")

    plt.xlabel("Frequency (MHz)")
    plt.ylabel(ylabel)
    plt.title(title)
    plt.grid(True, alpha=0.3)
    plt.legend()
    if force_zero_bottom:
        plt.ylim(bottom=0)
    plt.tight_layout()
    plt.savefig(out_png)
    plt.close()

def plot_stat_with_mask(
    stat_vec: np.ndarray,
    masked_bins: np.ndarray,
    out_png: str,
    title: str,
    ylabel: str,
) -> None:
    x = np.arange(stat_vec.size) * CHANNEL_BIN_MHZ
    y = np.asarray(stat_vec, dtype=float).copy()

    plt.figure(figsize=(10, 5.5))
    plt.plot(x, y, label="Statistic")

    if masked_bins.size > 0:
        mb = masked_bins[(masked_bins >= 0) & (masked_bins < stat_vec.size)]
        plt.scatter(x[mb], y[mb], s=18, label="Masked bins")

    plt.xlabel("Frequency (MHz)")
    plt.ylabel(ylabel)
    plt.title(title)
    plt.grid(True, alpha=0.3)
    plt.legend()
    plt.tight_layout()
    plt.savefig(out_png)
    plt.close()


def plot_eigenspectrum_first20(
    eigenvals: np.ndarray,
    out_png: str,
    title: str
) -> None:
    if eigenvals.size == 0:
        return

    n_plot = min(20, eigenvals.size)
    vals = eigenvals[:n_plot]
    idx = np.arange(1, n_plot + 1)

    eps = 1e-16
    vals_safe = np.where(vals > 0, vals, eps)

    plt.figure(figsize=(6, 4))
    plt.plot(idx, vals_safe, marker="o")
    plt.yscale("log")
    plt.xlabel("Principal component index")
    plt.ylabel("Eigenvalue (log scale)")
    plt.title(title + " (first 20 components)")
    plt.grid(True, which="both", alpha=0.3)
    plt.ylim(bottom=0)
    plt.tight_layout()
    plt.savefig(out_png)
    plt.close()


def plot_stacked_eigenvalues_first20(
    eigenvals_1: np.ndarray,
    eigenvals_2: Optional[np.ndarray],
    eigenvals_3: Optional[np.ndarray],
    out_pdf: Optional[PdfPages],
    out_png: str,
    title: str,
) -> None:
    """
    Plot up to the first 20 eigenvalues for SVD 1/2/3 on the same axes.
    Intended mainly for the auto-product pipeline where all three stages exist.
    """
    if eigenvals_1 is None or np.asarray(eigenvals_1).size == 0:
        return

    fig = plt.figure(figsize=(7.5, 5.0))
    idx1 = np.arange(1, min(20, np.asarray(eigenvals_1).size) + 1)
    vals1 = np.asarray(eigenvals_1[:idx1.size], dtype=float)
    vals1 = np.where(vals1 > 0, vals1, 1e-16)
    plt.plot(idx1, vals1, marker="o", label="SVD 1")

    if eigenvals_2 is not None and np.asarray(eigenvals_2).size > 0:
        idx2 = np.arange(1, min(20, np.asarray(eigenvals_2).size) + 1)
        vals2 = np.asarray(eigenvals_2[:idx2.size], dtype=float)
        vals2 = np.where(vals2 > 0, vals2, 1e-16)
        plt.plot(idx2, vals2, marker="o", label="SVD 2")

    if eigenvals_3 is not None and np.asarray(eigenvals_3).size > 0:
        idx3 = np.arange(1, min(20, np.asarray(eigenvals_3).size) + 1)
        vals3 = np.asarray(eigenvals_3[:idx3.size], dtype=float)
        vals3 = np.where(vals3 > 0, vals3, 1e-16)
        plt.plot(idx3, vals3, marker="o", label="SVD 3")

    plt.yscale("log")
    plt.xlabel("Principal component index")
    plt.ylabel("Eigenvalue (log scale)")
    plt.title(title + " (first 20 components)")
    plt.grid(True, which="both", alpha=0.3)
    plt.legend()
    plt.tight_layout()
    plt.savefig(out_png)
    if out_pdf is not None:
        out_pdf.savefig(fig)
    plt.close(fig)



def plot_residual_example(
    R: np.ndarray,
    out_png: str,
    title: str,
    ylabel: str,
    max_lines: int = 12
) -> None:
    n_samples, n_bins = R.shape
    x = np.arange(n_bins) * CHANNEL_BIN_MHZ

    plt.figure(figsize=(9, 5.5))
    nplot = min(n_samples, max_lines)
    for i in range(nplot):
        y = R[i, :].astype(float).copy()
        y[y == 0.0] = np.nan
        plt.plot(x, y, linewidth=1.0, alpha=0.9)

    plt.xlabel("Frequency (MHz)")
    plt.ylabel(ylabel)
    plt.title(title + f" (showing {nplot}/{n_samples} samples)")
    plt.grid(True, alpha=0.3)
    plt.tight_layout()
    plt.savefig(out_png)
    plt.close()


def plot_first5_eigenvectors_to_pdf(
    eigvecs: np.ndarray,
    title: str,
    pdf: PdfPages,
    max_components: int = 5,
    omitted_bins: Optional[np.ndarray] = None,
) -> None:
    """
    Append two pages to the eigenvector PDF:
      1) first few eigenvectors overplotted on one page
      2) EV1 and EV2 each on their own axis (stacked vertically)

    Omitted bins are removed from the plotted x/y arrays entirely so the
    eigenvector plots only show bins that actually participate in the SVD.
    """
    if eigvecs.ndim != 2 or eigvecs.size == 0:
        return

    n_bins, n_comp = eigvecs.shape
    keep = np.ones(n_bins, dtype=bool)
    if omitted_bins is not None and np.size(omitted_bins) > 0:
        omitted_bins = np.asarray(omitted_bins, dtype=int)
        valid = (omitted_bins >= 0) & (omitted_bins < n_bins)
        keep[omitted_bins[valid]] = False

    x = (np.arange(n_bins) * CHANNEL_BIN_MHZ)[keep]
    eigvecs_plot = eigvecs[keep, :]

    if x.size == 0 or eigvecs_plot.size == 0:
        return

    k = min(max_components, n_comp)

    fig = plt.figure(figsize=(10, 7))
    for i in range(k):
        plt.plot(x, eigvecs_plot[:, i], label=f"EV{i+1}", linewidth=1.2)

    ymax = float(np.max(np.abs(eigvecs_plot[:, :k]))) if k > 0 else 0.0
    if np.isfinite(ymax) and ymax > 0.0:
        plt.ylim(-1.1 * ymax, 1.1 * ymax)

    plt.xlabel("Frequency (MHz)")
    plt.ylabel("Eigenvector amplitude")
    plt.title(title)
    plt.grid(True, alpha=0.3)
    plt.legend()
    plt.tight_layout()

    pdf.savefig(fig)
    plt.close(fig)

    n_show = min(2, n_comp)
    if n_show > 0:
        fig, axes = plt.subplots(n_show, 1, figsize=(10, 7), sharex=True)
        if n_show == 1:
            axes = [axes]

        for i in range(n_show):
            axes[i].plot(x, eigvecs_plot[:, i], linewidth=1.2)
            ymax_i = float(np.max(np.abs(eigvecs_plot[:, i]))) if eigvecs_plot.shape[0] > 0 else 0.0
            if np.isfinite(ymax_i) and ymax_i > 0.0:
                axes[i].set_ylim(-1.1 * ymax_i, 1.1 * ymax_i)
            axes[i].set_ylabel(f"EV{i+1}")
            axes[i].grid(True, alpha=0.3)
            axes[i].set_title(f"EV{i+1}")

        axes[-1].set_xlabel("Frequency (MHz)")
        plt.tight_layout()
        pdf.savefig(fig)
        plt.close(fig)


def plot_mean_with_preamp_overlay(
    mean_original: np.ndarray,
    mean_preamp: np.ndarray,
    masked_bins: np.ndarray,
    out_png: str,
    title: str,
    ylabel: str = "Power / spectrometer units",
    force_zero_bottom: bool = False,
) -> None:
    x = np.arange(mean_original.size) * CHANNEL_BIN_MHZ

    y_orig = np.asarray(mean_original, dtype=float).copy()
    y_pre = np.asarray(mean_preamp, dtype=float).copy()

    if masked_bins is not None and masked_bins.size > 0:
        valid = (masked_bins >= 0) & (masked_bins < mean_original.size)
        mb = masked_bins[valid]
        y_orig[mb] = np.nan
        y_pre[mb] = np.nan

    plt.figure(figsize=(10, 5.5))
    plt.plot(x, y_orig, linewidth=1.3, label="Original mean spectrum")
    plt.plot(x, y_pre, linewidth=1.3, label="Mean modeled preamp noise")
    plt.xlabel("Frequency (MHz)")
    plt.ylabel(ylabel)
    plt.title(title)
    plt.grid(True, alpha=0.3)
    plt.legend()
    if force_zero_bottom:
        plt.ylim(bottom=0)
    plt.tight_layout()
    plt.savefig(out_png)
    plt.close()



def plot_mean_comparison(
    mean_a: np.ndarray,
    mean_b: np.ndarray,
    masked_bins: np.ndarray,
    out_png: str,
    title: str,
    label_a: str,
    label_b: str,
    ylabel: str = "Power",
    force_zero_bottom: bool = False,
) -> None:
    x = np.arange(mean_a.size) * CHANNEL_BIN_MHZ
    y_a = np.asarray(mean_a, dtype=float).copy()
    y_b = np.asarray(mean_b, dtype=float).copy()

    if masked_bins is not None and masked_bins.size > 0:
        valid = (masked_bins >= 0) & (masked_bins < mean_a.size)
        mb = masked_bins[valid]
        y_a[mb] = np.nan
        y_b[mb] = np.nan

    plt.figure(figsize=(10, 5.5))
    plt.plot(x, y_a, linewidth=1.3, label=label_a)
    plt.plot(x, y_b, linewidth=1.3, label=label_b)
    plt.xlabel("Frequency (MHz)")
    plt.ylabel(ylabel)
    plt.title(title)
    plt.grid(True, alpha=0.3)
    plt.legend()
    if force_zero_bottom:
        plt.ylim(bottom=0)
    plt.tight_layout()
    plt.savefig(out_png)
    plt.close()

# ----------------------------
# Telemetry / gain extraction
# ----------------------------

def telemetry_attr_name(key: str) -> str:
    return str(TELEMETRY_FIELD_ALIASES.get(key, key))


def _as_1d_float_array(x) -> np.ndarray:
    arr = np.asarray(x)
    if arr.ndim == 0:
        arr = arr.reshape(1)
    return np.asarray(arr).reshape(-1)


def extract_attr_series_or_nan(s, key: str, n_time: int) -> np.ndarray:
    """
    Extract a telemetry-like attribute from Spectra, reshaped to (n_time,).
    If missing or mis-sized, return NaNs.
    """
    attr = telemetry_attr_name(key)
    if not hasattr(s, attr):
        return np.full(n_time, np.nan, dtype=float)

    try:
        arr = _as_1d_float_array(getattr(s, attr)).astype(float)
    except Exception:
        return np.full(n_time, np.nan, dtype=float)

    if arr.size == n_time:
        return arr

    if arr.size == 1:
        return np.full(n_time, float(arr[0]), dtype=float)

    return np.full(n_time, np.nan, dtype=float)


def infer_gain_level_from_actual_gain(actual_gain_array: np.ndarray, channel: int) -> np.ndarray:
    """
    Map actual_gain per row for one auto channel to gain labels.

    Advisor-confirmed mapping:
      0 -> L
      1 -> M
      2 -> H

    Returns object array of shape (n_rows,), entries in {"L", "M", "H", ""}.
    """
    ag = np.asarray(actual_gain_array, dtype=float)
    if ag.ndim != 2 or ag.shape[1] != 4:
        raise ValueError(f"actual_gain must have shape (n_rows, 4); got {ag.shape}")

    out = np.full(ag.shape[0], "", dtype=object)
    vals = ag[:, channel]
    for i, v in enumerate(vals):
        if not np.isfinite(v):
            continue
        iv = int(v)
        if iv == 0:
            out[i] = "L"
        elif iv == 1:
            out[i] = "M"
        elif iv == 2:
            out[i] = "H"
    return out

def infer_gain_level_array_from_gain(gain_raw, n_time: int) -> np.ndarray:
    """
    Convert per-row gain array into L/M/H labels where possible.
    Supports:
      - strings like L, M, H, L0, M2, H3
      - numeric arrays with unique values in {0,1,2} -> L,M,H
      - numeric arrays with unique values in {1,2,3} -> L,M,H
    Returns object array of shape (n_time,), entries in {"L","M","H",""}.
    """
    out = np.full(n_time, "", dtype=object)

    if gain_raw is None:
        return out

    arr = np.asarray(gain_raw)
    if arr.ndim == 0:
        arr = np.full(n_time, arr)
    arr = arr.reshape(-1)

    if arr.size == 1:
        arr = np.full(n_time, arr[0])

    if arr.size != n_time:
        return out

    # String-like case
    if arr.dtype.kind in {"U", "S", "O"}:
        for i, val in enumerate(arr):
            s = str(val).strip().upper()
            lvl = ""
            for cand in ("L", "M", "H"):
                if cand in s:
                    lvl = cand
                    break
            out[i] = lvl
        return out

    # Numeric case
    try:
        arrf = arr.astype(float)
    except Exception:
        return out

    finite = np.isfinite(arrf)
    uniq = sorted(np.unique(arrf[finite]).tolist()) if np.any(finite) else []

    mapping: Dict[float, str] = {}
    if set(uniq).issubset({0.0, 1.0, 2.0}):
        mapping = {0.0: "L", 1.0: "M", 2.0: "H"}
    elif set(uniq).issubset({1.0, 2.0, 3.0}):
        mapping = {1.0: "L", 2.0: "M", 3.0: "H"}
    else:
        return out

    for i, val in enumerate(arrf):
        if np.isfinite(val) and val in mapping:
            out[i] = mapping[val]

    return out


def extract_session_model_inputs(s, n_time: int) -> Dict[str, np.ndarray]:
    """
    Extract all telemetry needed for the gain + preamp models plus actual_gain.
    """
    out: Dict[str, np.ndarray] = {}

    for key in ALL_REQUIRED_TELEMETRY:
        out[key] = extract_attr_series_or_nan(s, key, n_time)

    actual_gain_raw = getattr(s, "actual_gain", None)
    if actual_gain_raw is None:
        out["actual_gain"] = np.full((n_time, 4), np.nan, dtype=float)
    else:
        try:
            ag = np.asarray(actual_gain_raw, dtype=float)
            if ag.ndim == 1 and ag.size == 4:
                ag = np.tile(ag.reshape(1, 4), (n_time, 1))
            elif ag.ndim == 2 and ag.shape == (n_time, 4):
                pass
            else:
                ag = np.full((n_time, 4), np.nan, dtype=float)
            out["actual_gain"] = ag
        except Exception:
            out["actual_gain"] = np.full((n_time, 4), np.nan, dtype=float)

    time_raw = getattr(s, "time", None)
    if time_raw is None:
        out["time"] = np.arange(n_time, dtype=float)
    else:
        try:
            t = _as_1d_float_array(time_raw).astype(float)
            if t.size == n_time:
                out["time"] = t
            else:
                out["time"] = np.arange(n_time, dtype=float)
        except Exception:
            out["time"] = np.arange(n_time, dtype=float)

    return out


def subset_model_inputs(model_inputs: Dict[str, np.ndarray], keep_mask: np.ndarray) -> Dict[str, np.ndarray]:
    out: Dict[str, np.ndarray] = {}
    for k, v in model_inputs.items():
        arr = np.asarray(v)
        if arr.shape[0] == keep_mask.shape[0]:
            out[k] = arr[keep_mask]
        else:
            out[k] = arr
    return out


# ----------------------------
# Corrected-gain model helpers
# (adapted from get_corrected_gain.py)
# ----------------------------

def make_gain_name(level: str, channel: int) -> str:
    level = str(level).strip().upper()
    if level not in {"L", "M", "H"}:
        raise ValueError("level must be one of: L, M, H")
    if channel not in {0, 1, 2, 3}:
        raise ValueError("channel must be one of: 0, 1, 2, 3")
    return f"{level}{channel}"


def get_telemetry_cols_for_gain(gain: str) -> List[str]:
    ch = int(gain[-1])
    adc_col = "SPE_ADC0_T" if ch in (0, 1) else "SPE_ADC1_T"
    return ["THERM_FPGA", adc_col, "SPE_1VAD8_V", "VMON_1V2D", "SPE_1VAD8_C"]


def build_feature_matrix(X: np.ndarray, tele_cols: List[str], order: int = 2) -> Tuple[np.ndarray, List[str]]:
    n_samples, n_features = X.shape
    Z = np.ones((n_samples, 1), dtype=float)
    feature_labels = ["1"]

    if order >= 1 and n_features > 0:
        Z = np.hstack([Z, X.astype(float)])
        feature_labels += list(tele_cols)

    if order == 2:
        quad_blocks = []
        quad_labels = []

        idx_th = tele_cols.index("THERM_FPGA") if "THERM_FPGA" in tele_cols else None
        adc_idx, adc_name = None, None
        for cand in ("SPE_ADC0_T", "SPE_ADC1_T", "TADC"):
            if cand in tele_cols:
                adc_idx = tele_cols.index(cand)
                adc_name = cand
                break

        if idx_th is not None:
            th = X[:, idx_th]
            quad_blocks.append((th * th).reshape(-1, 1))
            quad_labels.append("THERM_FPGA*THERM_FPGA")

        if adc_idx is not None:
            adc = X[:, adc_idx]
            quad_blocks.append((adc * adc).reshape(-1, 1))
            quad_labels.append(f"{adc_name}*{adc_name}")

        if idx_th is not None and adc_idx is not None:
            th = X[:, idx_th]
            adc = X[:, adc_idx]
            quad_blocks.append((th * adc).reshape(-1, 1))
            quad_labels.append(f"THERM_FPGA*{adc_name}")

        if quad_blocks:
            Z = np.hstack([Z] + quad_blocks)
            feature_labels += quad_labels

    return Z, feature_labels


def build_single_feature_row(gain: str, telemetry_values: Dict[str, float]) -> Tuple[np.ndarray, List[str], List[str]]:
    tele_cols = get_telemetry_cols_for_gain(gain)

    missing = [c for c in tele_cols if c not in telemetry_values or telemetry_values[c] is None or not np.isfinite(telemetry_values[c])]
    if missing:
        raise ValueError(f"Missing telemetry inputs for {gain}: {missing}")

    X = np.array([[float(telemetry_values[c]) for c in tele_cols]], dtype=float)
    Z, labels = build_feature_matrix(X, tele_cols, order=2)

    return Z[0], labels, tele_cols


def predict_pc_from_refit(
    alphas_refit: "np.ndarray",
    gain: str,
    pc: str,
    feature_row: np.ndarray,
    labels: List[str]
) -> float:
    sub = alphas_refit[
        (alphas_refit["gain_setting"] == gain) &
        (alphas_refit["component"] == pc) &
        (alphas_refit["model"] == "quadratic")
    ]

    if sub.empty:
        return 0.0

    label_to_value = {lbl: feature_row[i] for i, lbl in enumerate(labels)}

    yhat = 0.0
    for _, row in sub.iterrows():
        term = str(row["term"])
        alpha = float(row["alpha_refit"])
        if term in label_to_value:
            yhat += alpha * label_to_value[term]

    return float(yhat)


def reconstruct_gain_from_pcs(mean_vec: np.ndarray, eigvecs: np.ndarray, pc1: float, pc2: float) -> np.ndarray:
    v1 = eigvecs[:, 0]
    v2 = eigvecs[:, 1]
    pred = mean_vec + pc1 * v1 + pc2 * v2
    return pred


class GainModelCache:
    def __init__(self, gain_model_root: str):
        self.gain_model_root = os.path.expanduser(gain_model_root)
        self.preamp_dir = os.path.join(self.gain_model_root, "corrected", "phase1", "pca")
        self.alphas_dir = os.path.join(self.gain_model_root, "corrected", "phase2", "alphas")
        self.alphas_refit = None
        self.per_gain: Dict[str, Dict[str, np.ndarray]] = {}

    def _load_alphas(self):
        if self.alphas_refit is None:
            alpha_path = os.path.join(self.alphas_dir, "alpha_refit.csv")
            if not os.path.exists(alpha_path):
                raise FileNotFoundError(f"Missing alpha_refit.csv: {alpha_path}")
            import pandas as pd
            self.alphas_refit = pd.read_csv(alpha_path)

    def load_gain_products(self, gain: str) -> Dict[str, np.ndarray]:
        if gain in self.per_gain:
            return self.per_gain[gain]

        self._load_alphas()

        mean_path = os.path.join(self.preamp_dir, f"{gain}_mean.npy")
        eig_path = os.path.join(self.preamp_dir, f"{gain}_eigvecs.npy")
        freq_path = os.path.join(self.preamp_dir, f"{gain}_freqs.npy")

        missing = [p for p in [mean_path, eig_path, freq_path] if not os.path.exists(p)]
        if missing:
            raise FileNotFoundError("Missing gain-model products:\n" + "\n".join(missing))

        mean_vec = np.load(mean_path)
        eigvecs = np.load(eig_path)
        freqs = np.load(freq_path)

        freqs = np.asarray(freqs, dtype=float).reshape(-1)
        mean_vec = np.asarray(mean_vec, dtype=float).reshape(-1)
        eigvecs = np.asarray(eigvecs, dtype=float)

        if eigvecs.ndim != 2 or eigvecs.shape[0] != mean_vec.size:
            raise ValueError(f"Bad eigvecs shape for {gain}: {eigvecs.shape}, mean size={mean_vec.size}")

        if freqs.size != mean_vec.size:
            raise ValueError(f"Frequency / mean length mismatch for {gain}: {freqs.size} vs {mean_vec.size}")

        order = np.argsort(freqs)
        freqs = freqs[order]
        mean_vec = mean_vec[order]
        eigvecs = eigvecs[order, :]

        self.per_gain[gain] = {
            "mean_vec": mean_vec,
            "eigvecs": eigvecs,
            "freqs": freqs,
            "alphas_refit": self.alphas_refit,
        }
        return self.per_gain[gain]

    def predict_gain_vector(
        self,
        level: str,
        channel: int,
        telemetry_values: Dict[str, float],
        freqs_mhz: np.ndarray,
    ) -> np.ndarray:
        gain = make_gain_name(level, channel)
        prod = self.load_gain_products(gain)

        feature_row, labels, _tele_cols = build_single_feature_row(gain, telemetry_values)

        pc1_hat = predict_pc_from_refit(prod["alphas_refit"], gain, "PC1", feature_row, labels)
        pc2_hat = predict_pc_from_refit(prod["alphas_refit"], gain, "PC2", feature_row, labels)

        anchor_gain = reconstruct_gain_from_pcs(prod["mean_vec"], prod["eigvecs"], pc1_hat, pc2_hat)

        spline = CubicSpline(prod["freqs"], anchor_gain, extrapolate=False)
        out = np.asarray(spline(freqs_mhz), dtype=float)
        return out


# ----------------------------
# Preamp-noise model helpers
# (adapted from get_preamp_noise.py)
# ----------------------------

FREQ_RE = re.compile(r"^\s*([0-9]*\.?[0-9]+)\s*([kKmMgG]?)\s*[hH][zZ]\s*$")


def parse_header_freq_to_hz(h: str) -> float:
    hs = str(h).strip()
    m = re.match(r"^\s*([0-9]*\.?[0-9]+)\s*([kKmMgG]?)\s*[mM]?[hH][zZ]\s*$", hs)
    if not m:
        raise ValueError(f"Could not parse frequency header: '{h}'")
    val = float(m.group(1))
    prefix = (m.group(2) or "").lower()
    mult = {"": 1.0, "k": 1e3, "m": 1e6, "g": 1e9}[prefix]
    return val * mult


def find_noise_anchor(ws) -> Tuple[int, int]:
    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            if isinstance(v, str) and v.strip().lower() == "noise":
                return cell.row, cell.column
    raise RuntimeError("Could not find a cell labeled 'Noise' in the sheet.")


def extract_thermal_noise_table(ws) -> Tuple[np.ndarray, np.ndarray, np.ndarray]:
    noise_row, _ = find_noise_anchor(ws)
    header_row = noise_row + 1

    headers = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(header_row, c).value
        if v is None:
            continue
        if isinstance(v, str):
            headers[c] = v.strip()

    temp_col = None
    freq_cols: List[Tuple[int, float]] = []

    for c, h in headers.items():
        if h.lower() == "temperature":
            temp_col = c
            continue
        if h.lower() == "std":
            continue
        try:
            f_hz = parse_header_freq_to_hz(h)
            freq_cols.append((c, f_hz))
        except Exception:
            pass

    if temp_col is None:
        raise RuntimeError("Could not find 'Temperature' column in the thermal Noise table.")
    if not freq_cols:
        raise RuntimeError("Could not find any frequency columns in the thermal Noise table.")

    freq_cols.sort(key=lambda x: x[1])
    F = np.array([f for _, f in freq_cols], dtype=float)

    T_list: List[float] = []
    Z_rows: List[List[float]] = []

    r = header_row + 1
    while r <= ws.max_row:
        t = ws.cell(r, temp_col).value
        if t is None or (isinstance(t, str) and t.strip() == ""):
            break
        if not isinstance(t, (int, float)):
            break

        row_vals = []
        ok = True
        for c, _f in freq_cols:
            v = ws.cell(r, c).value
            if v is None:
                ok = False
                break
            row_vals.append(float(v))

        if ok:
            T_list.append(float(t))
            Z_rows.append(row_vals)
        r += 1

    if len(T_list) < 3:
        raise RuntimeError("Not enough thermal rows to fit temperature model (need >= 3).")

    T = np.array(T_list, dtype=float)
    Z = np.array(Z_rows, dtype=float)

    order = np.argsort(T)
    T = T[order]
    Z = Z[order, :]

    return T, F, Z


def extract_dense_noise_sheet(ws) -> Tuple[np.ndarray, np.ndarray]:
    header_row = None
    freq_col = None
    noise_col = None

    max_scan_rows = min(ws.max_row, 50)

    for r in range(1, max_scan_rows + 1):
        row_vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        norm = [v.strip().lower() if isinstance(v, str) else None for v in row_vals]

        cand_freq = None
        cand_noise = None
        for c, v in enumerate(norm, start=1):
            if v is None:
                continue
            if cand_freq is None and "freq" in v:
                cand_freq = c
            if cand_noise is None and "noise" in v:
                cand_noise = c

        if cand_freq is not None and cand_noise is not None:
            header_row = r
            freq_col = cand_freq
            noise_col = cand_noise
            break

    if header_row is None or freq_col is None or noise_col is None:
        raise RuntimeError("Could not locate Frequency/Noise columns in 'Noise' sheet by header search.")

    F_list = []
    N_list = []

    r = header_row + 1
    while r <= ws.max_row:
        f = ws.cell(r, freq_col).value
        n = ws.cell(r, noise_col).value
        if f is None or n is None:
            break
        if not isinstance(f, (int, float)) or not isinstance(n, (int, float)):
            break
        F_list.append(float(f))
        N_list.append(float(n))
        r += 1

    if len(F_list) < 5:
        raise RuntimeError("Dense 'Noise' sheet did not yield enough rows (need >= 5).")

    F_dense = np.array(F_list, dtype=float)
    N_dense = np.array(N_list, dtype=float)

    order = np.argsort(F_dense)
    F_dense = F_dense[order]
    N_dense = N_dense[order]

    return F_dense, N_dense


def fit_centered_quadratic_per_frequency(
    T: np.ndarray,
    F_anchor: np.ndarray,
    Z: np.ndarray,
    T0: float
) -> Tuple[np.ndarray, np.ndarray, np.ndarray]:
    u = T - T0
    X = np.vstack([np.ones_like(u), u, u**2]).T

    A_arr = np.zeros(len(F_anchor), dtype=float)
    B_arr = np.zeros(len(F_anchor), dtype=float)
    C_arr = np.zeros(len(F_anchor), dtype=float)

    for j in range(len(F_anchor)):
        y = Z[:, j].astype(float)
        coeffs, *_ = np.linalg.lstsq(X, y, rcond=None)
        A_arr[j], B_arr[j], C_arr[j] = coeffs

    return A_arr, B_arr, C_arr



class PreampNoiseModel:
    def __init__(self, xlsx_path: str, thermal_sheet: str = "Thermal test", noise_sheet: str = "Noise"):
        self.xlsx_path = os.path.expanduser(xlsx_path)
        self.thermal_sheet = thermal_sheet
        self.noise_sheet = noise_sheet

        if not os.path.exists(self.xlsx_path):
            raise FileNotFoundError(f"Preamp xlsx not found: {self.xlsx_path}")

        wb = openpyxl.load_workbook(self.xlsx_path, data_only=True)

        if self.thermal_sheet not in wb.sheetnames:
            raise RuntimeError(f"Thermal sheet '{self.thermal_sheet}' not found in {self.xlsx_path}")
        if self.noise_sheet not in wb.sheetnames:
            raise RuntimeError(f"Noise sheet '{self.noise_sheet}' not found in {self.xlsx_path}")

        ws_th = wb[self.thermal_sheet]
        ws_ns = wb[self.noise_sheet]

        self.T, self.F_anchor, self.Z = extract_thermal_noise_table(ws_th)
        self.F_dense, self.N_dense = extract_dense_noise_sheet(ws_ns)

        if not (np.min(self.T) <= T0_C <= np.max(self.T)):
            raise ValueError(
                f"T0={T0_C}°C is outside thermal temperature range "
                f"[{np.min(self.T)}, {np.max(self.T)}] for {self.xlsx_path}"
            )

        self.A_arr, self.B_arr, self.C_arr = fit_centered_quadratic_per_frequency(
            self.T, self.F_anchor, self.Z, T0_C
        )

        self.logF_anchor = np.log10(self.F_anchor)
        self.logF_dense = np.log10(self.F_dense)
        self.base_spline = InterpolatedUnivariateSpline(
            self.logF_dense,
            self.N_dense,
            k=min(3, len(self.F_dense) - 1),
        )

        self.temp_min = float(np.min(self.T))
        self.temp_max = float(np.max(self.T))
        self.freq_anchor_min = float(np.min(self.F_anchor))
        self.freq_anchor_max = float(np.max(self.F_anchor))
        self.freq_dense_min = float(np.min(self.F_dense))
        self.freq_dense_max = float(np.max(self.F_dense))
        self.freq_valid_min = max(self.freq_anchor_min, self.freq_dense_min)
        self.freq_valid_max = min(self.freq_anchor_max, self.freq_dense_max)

        if self.freq_valid_min >= self.freq_valid_max:
            raise ValueError(
                f"No overlapping valid frequency range for {self.xlsx_path}: "
                f"anchor=[{self.freq_anchor_min}, {self.freq_anchor_max}], "
                f"dense=[{self.freq_dense_min}, {self.freq_dense_max}]"
            )

    def is_temperature_supported(self, temperature_c: float) -> bool:
        return self.temp_min <= float(temperature_c) <= self.temp_max

    def valid_frequency_mask(self, freqs_hz: np.ndarray) -> np.ndarray:
        freqs_hz = np.asarray(freqs_hz, dtype=float)
        return (
            np.isfinite(freqs_hz)
            & (freqs_hz > 0.0)
            & (freqs_hz >= self.freq_valid_min)
            & (freqs_hz <= self.freq_valid_max)
        )

    def predict_noise_vector(self, temperature_c: float, freqs_hz: np.ndarray) -> np.ndarray:
        freqs_hz = np.asarray(freqs_hz, dtype=float)
        out = np.full(freqs_hz.shape, np.nan, dtype=float)

        valid_freq = self.valid_frequency_mask(freqs_hz)
        if not np.any(valid_freq):
            return out

        if not self.is_temperature_supported(temperature_c):
            return out

        u_q = float(temperature_c) - T0_C
        N_th_Tq = self.A_arr + self.B_arr * u_q + self.C_arr * (u_q ** 2)
        N_th_T0 = self.A_arr

        good_anchor = np.isfinite(N_th_Tq) & np.isfinite(N_th_T0) & (N_th_T0 > 0.0)
        if np.count_nonzero(good_anchor) < 2:
            return out

        scale_at_anchor = N_th_Tq[good_anchor] / N_th_T0[good_anchor]
        logF_anchor_good = self.logF_anchor[good_anchor]
        if len(logF_anchor_good) < 2:
            return out

        k_scale = min(3, len(logF_anchor_good) - 1)
        scale_spline = InterpolatedUnivariateSpline(
            logF_anchor_good,
            scale_at_anchor,
            k=k_scale,
        )

        logf = np.log10(freqs_hz[valid_freq])
        S = scale_spline(logf)
        N_base = self.base_spline(logf)
        modeled = N_base * S

        modeled[~np.isfinite(modeled)] = np.nan
        modeled[modeled < 0.0] = np.nan

        out[valid_freq] = modeled
        return out


# ----------------------------
# Notch-based grouping helpers
# ----------------------------
# Notch-based grouping helpers
# ----------------------------

def inspect_file_for_phase_assignment(path: str, session_dir_root: str) -> Optional[Dict[str, object]]:
    campaign_name = extract_campaign_name(path, session_dir_root)
    campaign_num = extract_campaign_number(path, session_dir_root)
    session_num = extract_session_number(path)

    try:
        s = Spectra(path)
        data = np.asarray(s.data)
    except Exception as e:
        print(f"[WARN] Skipping {campaign_name}/session {session_num:03d} ({path}) because it could not be loaded: {e}")
        return None

    if data.ndim != 3:
        print(
            f"[WARN] Skipping {campaign_name}/session {session_num:03d} ({path}) because "
            f"s.data has shape {data.shape}, expected (time, product, bin)"
        )
        return None

    n_time, n_prod, n_bins = data.shape
    if n_time == 0 or n_prod == 0 or n_bins == 0:
        print(
            f"[WARN] Skipping {campaign_name}/session {session_num:03d} ({path}) because "
            f"it has an empty dimension: shape={data.shape}"
        )
        return None

    try:
        notch = np.asarray(getattr(s, "notch", []))
        has_notch_zero = bool(notch.size > 0 and np.any(notch == 0))
        notch_unique = np.unique(notch).tolist() if notch.size else []
    except Exception:
        has_notch_zero = False
        notch_unique = []

    return {
        "campaign_name": campaign_name,
        "campaign_num": campaign_num,
        "session_num": session_num,
        "path": path,
        "has_notch_zero": has_notch_zero,
        "notch_unique": notch_unique,
        "shape": (n_time, n_prod, n_bins),
    }


def assign_phases_by_notch(
    paths: List[str],
    session_dir_root: str
) -> Tuple[Dict[int, List[Tuple[int, str]]], List[Dict[str, object]]]:
    inspected: List[Dict[str, object]] = []
    for path in paths:
        try:
            info = inspect_file_for_phase_assignment(path, session_dir_root)
        except Exception as e:
            print(f"[WARN] Could not inspect {path}: {e}")
            info = None
        if info is not None:
            inspected.append(info)

    inspected.sort(key=lambda d: (int(d["campaign_num"]), int(d["session_num"]), str(d["path"])))

    groups: Dict[int, List[Tuple[int, str]]] = defaultdict(list)
    assignment_rows: List[Dict[str, object]] = []

    current_phase: Optional[int] = None

    for info in inspected:
        path = str(info["path"])
        session_num = int(info["session_num"])
        campaign_name = str(info["campaign_name"])
        has_notch_zero = bool(info["has_notch_zero"])

        if has_notch_zero:
            current_phase = 0
            assignment_rows.append({
                "campaign_name": campaign_name,
                "campaign_num": int(info["campaign_num"]),
                "session_num": session_num,
                "path": path,
                "has_notch_zero": has_notch_zero,
                "assigned_phase": 0,
                "status": "anchor_only",
                "reason": "notch==0 reset",
                "shape": info["shape"],
                "notch_unique": info["notch_unique"],
            })
            print(
                f"[Info] Registered anchor {campaign_name}/session {session_num:03d} "
                f"(notch_zero=True); nmod4=0 not analyzed"
            )
            continue

        if current_phase is None:
            print(f"[WARN] No prior notch==0 anchor before {campaign_name}/session {session_num:03d}; skipping")
            assignment_rows.append({
                "campaign_name": campaign_name,
                "campaign_num": int(info["campaign_num"]),
                "session_num": session_num,
                "path": path,
                "has_notch_zero": has_notch_zero,
                "assigned_phase": "",
                "status": "skipped_before_first_anchor",
                "reason": "No prior notch==0 anchor",
                "shape": info["shape"],
                "notch_unique": info["notch_unique"],
            })
            continue

        phase = (current_phase + 1) % 4
        current_phase = phase

        if phase == 0:
            assignment_rows.append({
                "campaign_name": campaign_name,
                "campaign_num": int(info["campaign_num"]),
                "session_num": session_num,
                "path": path,
                "has_notch_zero": has_notch_zero,
                "assigned_phase": phase,
                "status": "anchor_only",
                "reason": "wrapped_to_nmod4_0_not_analyzed",
                "shape": info["shape"],
                "notch_unique": info["notch_unique"],
            })
            print(
                f"[Info] Registered anchor {campaign_name}/session {session_num:03d} "
                f"(notch_zero=False); nmod4=0 not analyzed"
            )
            continue

        groups[phase].append((session_num, path))
        assignment_rows.append({
            "campaign_name": campaign_name,
            "campaign_num": int(info["campaign_num"]),
            "session_num": session_num,
            "path": path,
            "has_notch_zero": has_notch_zero,
            "assigned_phase": phase,
            "status": "used",
            "reason": "advanced_from_previous",
            "shape": info["shape"],
            "notch_unique": info["notch_unique"],
        })

        print(
            f"[Info] Assigned {campaign_name}/session {session_num:03d} "
            f"(notch_zero={has_notch_zero}) -> nmod4={phase}"
        )

    return dict(groups), assignment_rows


# ----------------------------
# Full data loading for a group
# ----------------------------

def load_group_data(
    group_items: List[Tuple[int, str]],
    session_dir_root: str
) -> Tuple[List[Dict[str, object]], int, int]:
    sessions_info: List[Dict[str, object]] = []
    ref_n_prod = None
    ref_n_bins = None

    for session_num, path in group_items:
        campaign_name = extract_campaign_name(path, session_dir_root)

        try:
            s = Spectra(path)
            data = np.asarray(s.data, dtype=float)

            if data.ndim != 3:
                raise ValueError(f"s.data has shape {data.shape}, expected (time, product, bin)")

            _n_time_tmp, _n_prod_tmp, _n_bins_tmp = data.shape
            bitslice_rows, bitslice_scale_rows = get_rowwise_bitslice_scale_from_h5(
                path,
                expected_n_rows=_n_time_tmp,
                n_prod=_n_prod_tmp,
            )
            data = data * bitslice_scale_rows[:, :, None]
        except Exception as e:
            print(f"[WARN] Skipping {campaign_name}/session {session_num:03d} ({path}) because it could not be loaded/scaled: {e}")
            continue

        if data.ndim != 3:
            print(
                f"[WARN] Skipping {campaign_name}/session {session_num:03d} ({path}) because "
                f"s.data has shape {data.shape}, expected (time, product, bin)"
            )
            continue

        n_time, n_prod, n_bins = data.shape

        if n_time == 0 or n_prod == 0 or n_bins == 0:
            print(
                f"[WARN] Skipping {campaign_name}/session {session_num:03d} ({path}) because "
                f"it has an empty dimension: shape={data.shape}"
            )
            continue

        if ref_n_prod is None:
            ref_n_prod = n_prod
            ref_n_bins = n_bins
        else:
            if n_prod != ref_n_prod or n_bins != ref_n_bins:
                print(
                    f"[WARN] Skipping {campaign_name}/session {session_num:03d} ({path}) because "
                    f"its shape ({n_time}, {n_prod}, {n_bins}) does not match "
                    f"the group reference (_, {ref_n_prod}, {ref_n_bins})"
                )
                continue

        try:
            g = np.asarray(getattr(s, "gain", []))
            guniq = np.unique(g) if g.size else np.array([])
        except Exception:
            guniq = np.array([])

        try:
            ag = np.asarray(getattr(s, "actual_gain", []))
            aguniq = np.unique(ag) if ag.size else np.array([])
        except Exception:
            aguniq = np.array([])

        try:
            notch = np.asarray(getattr(s, "notch", []))
            notch_u = np.unique(notch) if notch.size else np.array([])
        except Exception:
            notch_u = np.array([])

        bitslice_unique = np.unique(bitslice_rows) if bitslice_rows.size else np.array([])
        bitslice_scale_unique = np.unique(bitslice_scale_rows) if bitslice_scale_rows.size else np.array([])

        model_inputs = extract_session_model_inputs(s, n_time)

        sessions_info.append(
            {
                "campaign_name": campaign_name,
                "session_num": session_num,
                "path": path,
                "data": data,
                "n_time": n_time,
                "n_prod": n_prod,
                "n_bins": n_bins,
                "gain_unique": guniq,
                "actual_gain_unique": aguniq,
                "notch_unique": notch_u,
                "bitslice_values": bitslice_rows,
                "bitslice_scale": bitslice_scale_rows,
                "bitslice_unique": bitslice_unique,
                "bitslice_scale_unique": bitslice_scale_unique,
                "model_inputs": model_inputs,
            }
        )

    if not sessions_info:
        raise RuntimeError("No valid session files remained in this nmod4 group after filtering")

    return sessions_info, ref_n_prod, ref_n_bins


# ----------------------------
# Preamp modeling for auto products
# ----------------------------

def infer_cross_product_gain_pair(actual_gain_array: np.ndarray, prod: int) -> np.ndarray:
    """
    Return ordered gain-pair labels for cross-correlation products.
    Examples: LL, LM, LH, ML, MM, MH, HL, HM, HH.
    """
    if prod not in CROSS_PRODUCT_CHANNELS:
        raise ValueError(f"Product {prod} is not a cross-correlation product")

    ch0, ch1 = CROSS_PRODUCT_CHANNELS[prod]
    g0 = infer_gain_level_from_actual_gain(actual_gain_array, ch0)
    g1 = infer_gain_level_from_actual_gain(actual_gain_array, ch1)

    out = np.full(g0.shape[0], "", dtype=object)
    for i, (a, b) in enumerate(zip(g0, g1)):
        if a in {"L", "M", "H"} and b in {"L", "M", "H"}:
            out[i] = a + b
    return out


def collect_stacked_model_inputs_for_product(
    sessions_info: List[Dict[str, object]],
    X_row_count_expected: int,
    product: int,
) -> Dict[str, np.ndarray]:
    """
    Build per-row stacked arrays aligned with the original stacked X rows.
    Also infer gain label for the requested product:
      - products 0-3  -> L / M / H
      - products 4-15 -> LL / LM / ... / HH
    """
    parts: Dict[str, List[np.ndarray]] = defaultdict(list)

    for si in sessions_info:
        mi = si["model_inputs"]
        n_time = int(si["n_time"])
        for key, arr in mi.items():
            arr_np = np.asarray(arr)
            if key == "actual_gain":
                if arr_np.shape != (n_time, 4):
                    raise ValueError(f"Model input 'actual_gain' shape mismatch in session {si['path']}: {arr_np.shape}")
            else:
                if arr_np.shape[0] != n_time:
                    raise ValueError(f"Model input '{key}' length mismatch in session {si['path']}")
            parts[key].append(arr_np)

    out: Dict[str, np.ndarray] = {}
    for key, lst in parts.items():
        out[key] = np.concatenate(lst, axis=0)

    if "actual_gain" not in out:
        out["actual_gain"] = np.full((X_row_count_expected, 4), np.nan, dtype=float)

    if product <= 3:
        out["gain_level"] = infer_gain_level_from_actual_gain(out["actual_gain"], product)
    else:
        out["gain_level"] = infer_cross_product_gain_pair(out["actual_gain"], product)

    if out["gain_level"].shape[0] != X_row_count_expected:
        raise ValueError("Stacked model inputs do not align with stacked X rows")

    return out


def gain_groups_for_product(product: int) -> List[str]:
    if product <= 3:
        return ["L", "M", "H"]
    return ["LL", "LM", "LH", "ML", "MM", "MH", "HL", "HM", "HH"]

def write_model_input_summary_csv(
    path: str,
    row_meta: List[List[object]],
    model_inputs: Dict[str, np.ndarray],
    channel: int,
) -> None:
    header = [
        "sample",
        "campaign_name",
        "session_num",
        "session_name",
        "row_in_session",
        "gain_level",
        "actual_gain_ch0",
        "actual_gain_ch1",
        "actual_gain_ch2",
        "actual_gain_ch3",
        "THERM_FPGA",
        "SPE_ADC0_T",
        "SPE_ADC1_T",
        "SPE_1VAD8_V",
        "VMON_1V2D",
        "SPE_1VAD8_C",
        f"PFPS_PA{channel}_T",
    ]

    pfps_key = f"PFPS_PA{channel}_T"
    ag = np.asarray(model_inputs["actual_gain"], dtype=float)
    rows = []
    for i, meta in enumerate(row_meta):
        rows.append(
            meta + [
                model_inputs["gain_level"][i],
                ag[i, 0], ag[i, 1], ag[i, 2], ag[i, 3],
                model_inputs["THERM_FPGA"][i],
                model_inputs["SPE_ADC0_T"][i],
                model_inputs["SPE_ADC1_T"][i],
                model_inputs["SPE_1VAD8_V"][i],
                model_inputs["VMON_1V2D"][i],
                model_inputs["SPE_1VAD8_C"][i],
                model_inputs[pfps_key][i],
            ]
        )

    write_csv_rows(path, header, rows)


def extract_gain_telemetry_values(model_inputs: Dict[str, np.ndarray], row_idx: int) -> Dict[str, float]:
    return {
        "THERM_FPGA": float(model_inputs["THERM_FPGA"][row_idx]),
        "SPE_ADC0_T": float(model_inputs["SPE_ADC0_T"][row_idx]),
        "SPE_ADC1_T": float(model_inputs["SPE_ADC1_T"][row_idx]),
        "SPE_1VAD8_V": float(model_inputs["SPE_1VAD8_V"][row_idx]),
        "VMON_1V2D": float(model_inputs["VMON_1V2D"][row_idx]),
        "SPE_1VAD8_C": float(model_inputs["SPE_1VAD8_C"][row_idx]),
    }


def build_nv_sqrt_hz_matrix_for_gain_group(
    X: np.ndarray,
    product: int,
    gain_level: str,
    model_inputs: Dict[str, np.ndarray],
    bin_freqs_mhz: np.ndarray,
    gain_cache: GainModelCache,
) -> Tuple[np.ndarray, List[str], np.ndarray, np.ndarray]:
    n_rows, n_bins = X.shape
    out = np.zeros((n_rows, n_bins), dtype=float)
    warnings_list: List[str] = []
    row_keep_for_svd = np.ones(n_rows, dtype=bool)
    invalid_bins_mask = np.zeros(n_bins, dtype=bool)

    gain_level = str(gain_level).strip().upper()

    if product <= 3:
        channels = (product,)
        levels = (gain_level,)
    else:
        if len(gain_level) != 2:
            raise ValueError(f"Cross-product gain level must have length 2, got '{gain_level}'")
        channels = CROSS_PRODUCT_CHANNELS[product]
        levels = (gain_level[0], gain_level[1])

    n_converted = 0
    n_skipped = 0
    n_rows_without_any_valid_bins = 0

    for i in range(n_rows):
        try:
            telemetry_values = extract_gain_telemetry_values(model_inputs, i)
        except Exception as e:
            n_skipped += 1
            row_keep_for_svd[i] = False
            warnings_list.append(f"row {i}: failed to extract telemetry: {e}")
            continue

        needed = [
            telemetry_values["THERM_FPGA"],
            telemetry_values["SPE_ADC0_T"],
            telemetry_values["SPE_ADC1_T"],
            telemetry_values["SPE_1VAD8_V"],
            telemetry_values["VMON_1V2D"],
            telemetry_values["SPE_1VAD8_C"],
        ]
        if not np.all(np.isfinite(needed)):
            n_skipped += 1
            row_keep_for_svd[i] = False
            warnings_list.append(f"row {i}: missing/non-finite telemetry for gain conversion")
            continue

        try:
            gain_vectors = [
                gain_cache.predict_gain_vector(
                    level=level,
                    channel=channel,
                    telemetry_values=telemetry_values,
                    freqs_mhz=bin_freqs_mhz,
                )
                for level, channel in zip(levels, channels)
            ]
        except Exception as e:
            n_skipped += 1
            row_keep_for_svd[i] = False
            warnings_list.append(f"row {i}: failed to predict gain vector(s): {e}")
            continue

        gain_vectors = [np.asarray(gv, dtype=float) for gv in gain_vectors]

        if product <= 3:
            gain_denom = gain_vectors[0]
            valid = (
                np.isfinite(X[i, :])
                & np.isfinite(gain_denom)
                & (gain_denom > 0.0)
                & (X[i, :] >= 0.0)
            )

            invalid_bins_mask |= ~valid

            if not np.any(valid):
                n_rows_without_any_valid_bins += 1
                row_keep_for_svd[i] = False
                warnings_list.append(f"row {i}: no valid auto bins after gain conversion")
                continue

            row_out = np.zeros(n_bins, dtype=float)
            row_out[valid] = np.sqrt(X[i, valid] / gain_denom[valid])
            out[i, :] = row_out

        else:
            gain_geom = np.sqrt(gain_vectors[0] * gain_vectors[1])
            noise_row = np.asarray(X[i, :], dtype=float)

            valid = (
                np.isfinite(noise_row)
                & np.isfinite(gain_geom)
                & (gain_geom > 0.0)
            )

            invalid_bins_mask |= ~valid

            if not np.any(valid):
                n_rows_without_any_valid_bins += 1
                row_keep_for_svd[i] = False
                warnings_list.append(f"row {i}: no valid cross bins after gain conversion")
                continue

            row_out = np.zeros(n_bins, dtype=float)
            row_out[valid] = (
                np.sqrt(np.abs(noise_row[valid]) / gain_geom[valid])
                * np.sign(noise_row[valid])
            )
            out[i, :] = row_out

        n_converted += 1

    warnings_list.append(f"rows converted successfully: {n_converted}/{n_rows}")
    warnings_list.append(f"rows skipped entirely: {n_skipped}/{n_rows}")
    warnings_list.append(f"rows with no valid converted bins: {n_rows_without_any_valid_bins}/{n_rows}")

    return out, warnings_list, row_keep_for_svd, np.where(invalid_bins_mask)[0]
def build_preamp_noise_matrix_for_auto_product(
    channel: int,
    model_inputs: Dict[str, np.ndarray],
    bin_freqs_mhz: np.ndarray,
    preamp_models: Dict[int, PreampNoiseModel],
) -> Tuple[np.ndarray, List[str], np.ndarray, np.ndarray]:
    n_rows = np.asarray(model_inputs["gain_level"]).shape[0]
    n_bins = len(bin_freqs_mhz)
    out = np.zeros((n_rows, n_bins), dtype=float)
    warnings_list: List[str] = []

    freqs_hz = np.asarray(bin_freqs_mhz, dtype=float) * 1e6
    preamp_model = preamp_models[channel]
    pfps_key = f"PFPS_PA{channel}_T"

    row_keep_for_second_svd = np.ones(n_rows, dtype=bool)
    valid_freq_mask_preamp = preamp_model.valid_frequency_mask(freqs_hz)

    n_modeled = 0
    n_skipped = 0
    n_temp_unsupported = 0
    n_rows_without_any_valid_bins = 0

    for i in range(n_rows):
        pa_temp = float(model_inputs[pfps_key][i])
        if not np.isfinite(pa_temp):
            n_skipped += 1
            row_keep_for_second_svd[i] = False
            warnings_list.append(f"row {i}: missing/non-finite preamp temperature")
            continue

        if not preamp_model.is_temperature_supported(pa_temp):
            n_skipped += 1
            n_temp_unsupported += 1
            row_keep_for_second_svd[i] = False
            warnings_list.append(
                f"row {i}: preamp temperature {pa_temp:.6g} C outside supported range "
                f"[{preamp_model.temp_min:.6g}, {preamp_model.temp_max:.6g}]"
            )
            continue

        try:
            preamp_phys = preamp_model.predict_noise_vector(pa_temp, freqs_hz)
        except Exception as e:
            n_skipped += 1
            row_keep_for_second_svd[i] = False
            warnings_list.append(f"row {i}: preamp-noise model failed: {e}")
            continue

        model_row = np.zeros(n_bins, dtype=float)
        good = (
            valid_freq_mask_preamp
            & np.isfinite(preamp_phys)
            & (preamp_phys >= 0.0)
        )

        if not np.any(good):
            n_skipped += 1
            n_rows_without_any_valid_bins += 1
            row_keep_for_second_svd[i] = False
            warnings_list.append(
                f"row {i}: no valid preamp bins after range/positivity safeguards"
            )
            continue

        model_row[good] = preamp_phys[good]
        out[i, :] = model_row
        n_modeled += 1

        if np.any(~valid_freq_mask_preamp):
            warnings_list.append(
                f"row {i}: some bins excluded because frequency was outside supported preamp range"
            )
        if np.any(np.isfinite(preamp_phys) & (preamp_phys < 0.0) & valid_freq_mask_preamp):
            warnings_list.append(f"row {i}: some bins excluded because modeled preamp noise < 0")

    invalid_freq_bins = np.where(~valid_freq_mask_preamp)[0]

    warnings_list.append(f"modeled_rows={n_modeled}")
    warnings_list.append(f"skipped_rows={n_skipped}")
    warnings_list.append(f"rows_skipped_temperature_out_of_range={n_temp_unsupported}")
    warnings_list.append(f"rows_skipped_no_valid_bins={n_rows_without_any_valid_bins}")
    warnings_list.append(f"invalid_preamp_frequency_bins={len(invalid_freq_bins)}")
    warnings_list.append(
        f"valid_preamp_temperature_range_C=[{preamp_model.temp_min:.6g}, {preamp_model.temp_max:.6g}]"
    )
    warnings_list.append(
        f"valid_preamp_frequency_range_Hz=[{preamp_model.freq_valid_min:.6g}, {preamp_model.freq_valid_max:.6g}]"
    )

    return out, warnings_list, row_keep_for_second_svd, invalid_freq_bins


def save_preamp_model_summary(
    path: str,
    warnings_list: List[str],
) -> None:
    with open(path, "w") as f:
        for line in warnings_list:
            f.write(str(line) + "\n")


# ----------------------------
# Helpers for saving SVD results
# ----------------------------

def save_svd_products(
    X_for_mean: np.ndarray,
    R: np.ndarray,
    mu: np.ndarray,
    pcs: np.ndarray,
    eigvecs: np.ndarray,
    eigenvals: np.ndarray,
    row_meta: List[List[object]],
    prod_out: str,
    filename_prefix: str,
    title_prefix: str,
    residual_mode: str,
    mean_pdf: Optional[PdfPages],
    eigvec_pdf: Optional[PdfPages],
    omitted_bins: Optional[np.ndarray] = None,
    mean_ylabel: str = "Mean power",
    residual_ylabel: Optional[str] = None,
    force_zero_bottom: bool = False,
) -> None:
    np.save(os.path.join(prod_out, f"{filename_prefix}mean.npy"), mu)

    plot_mean_spectrum(
        mean_vec=mu,
        out_png=os.path.join(prod_out, f"{filename_prefix}mean.png"),
        title=short_stage_title(title_prefix, "Mean"),
        pdf=mean_pdf,
        ylabel=mean_ylabel,
        force_zero_bottom=force_zero_bottom,
    )

    if residual_ylabel is None:
        residual_ylabel = (
            "Frac resid"
            if residual_mode == "fractional"
            else "Abs resid"
        )

    plot_residual_example(
        R,
        out_png=os.path.join(prod_out, f"{filename_prefix}{residual_mode}_residual_example.png"),
        title=short_stage_title(title_prefix, "Residuals"),
        ylabel=residual_ylabel,
        max_lines=12,
    )

    np.save(os.path.join(prod_out, f"{filename_prefix}eigvecs.npy"), eigvecs)

    if eigvec_pdf is not None:
        plot_first5_eigenvectors_to_pdf(
            eigvecs=eigvecs,
            title=short_stage_title(title_prefix, "Eigvecs"),
            pdf=eigvec_pdf,
            max_components=5,
            omitted_bins=omitted_bins,
        )

    total = float(np.sum(eigenvals)) if eigenvals.size else 0.0
    eig_rows = []
    for i, lam in enumerate(eigenvals, start=1):
        ratio = float(lam / total) if total > 0 else 0.0
        eig_rows.append([i, float(lam), ratio])

    write_csv_rows(
        os.path.join(prod_out, f"{filename_prefix}eigenvalues.csv"),
        ["component", "eigenvalue", "explained_variance_ratio"],
        eig_rows,
    )

    plot_eigenspectrum_first20(
        eigenvals,
        out_png=os.path.join(prod_out, f"{filename_prefix}eigenspectrum_log_first20.png"),
        title=short_stage_title(title_prefix, "Eigvals"),
    )

    write_csv_rows(
        os.path.join(prod_out, f"{filename_prefix}sample_index_map.csv"),
        ["sample", "campaign_name", "session_num", "session_name", "row_in_session"],
        row_meta,
    )

    pc_header = ["sample", "campaign_name", "session_num", "session_name", "row_in_session"] + [
        f"PC{i+1}" for i in range(pcs.shape[1])
    ]
    pc_rows = []
    for i in range(pcs.shape[0]):
        meta_prefix = row_meta[i]
        pc_rows.append(meta_prefix + pcs[i, :].tolist())

    write_csv_rows(os.path.join(prod_out, f"{filename_prefix}pcs.csv"), pc_header, pc_rows)

    k5 = min(5, pcs.shape[1])
    pc5_header = ["sample", "campaign_name", "session_num", "session_name", "row_in_session"] + [
        f"PC{i+1}" for i in range(k5)
    ]
    pc5_rows = []
    for i in range(pcs.shape[0]):
        meta_prefix = row_meta[i]
        pc5_rows.append(meta_prefix + pcs[i, :k5].tolist())

    write_csv_rows(os.path.join(prod_out, f"{filename_prefix}pcs_first5.csv"), pc5_header, pc5_rows)



def build_lowfreq_repaired_matrix_for_third_svd(
    X_preamp_subtracted_raw: np.ndarray,
    bin_freqs_mhz: np.ndarray,
    original_mask_bins: np.ndarray,
    invalid_preamp_freq_bins: np.ndarray,
    fit_hi_mhz: float = 3.0,
    transition_hi_mhz: float = 5.0,
    preamp_low_cut_mhz: float = 0.3,
) -> Tuple[np.ndarray, np.ndarray, np.ndarray, List[str]]:
    """Repair the low-frequency end after preamp subtraction and before a third SVD."""
    X = np.asarray(X_preamp_subtracted_raw, dtype=float)
    freqs = np.asarray(bin_freqs_mhz, dtype=float).reshape(-1)
    if X.ndim != 2:
        raise ValueError(f"X_preamp_subtracted_raw must be 2D, got {X.shape}")
    if X.shape[1] != freqs.size:
        raise ValueError(f"Frequency length mismatch: X has {X.shape[1]} bins, freqs has {freqs.size}")

    invalid_preamp_freq_bins = np.asarray(invalid_preamp_freq_bins, dtype=int)
    original_mask_bins = np.asarray(original_mask_bins, dtype=int)

    valid_invalid = invalid_preamp_freq_bins[(invalid_preamp_freq_bins >= 0) & (invalid_preamp_freq_bins < freqs.size)]
    high_invalid_bins = valid_invalid[freqs[valid_invalid] >= preamp_low_cut_mhz]
    third_svd_mask_bins = np.union1d(original_mask_bins, high_invalid_bins)

    fit_bins = np.where((freqs >= preamp_low_cut_mhz) & (freqs < fit_hi_mhz))[0]
    transition_bins = np.where((freqs >= fit_hi_mhz) & (freqs < transition_hi_mhz))[0]
    low_fill_bins = np.where(freqs < fit_hi_mhz)[0]

    original_mask_set = set(int(b) for b in original_mask_bins.tolist())
    invalid_mask_set = set(int(b) for b in valid_invalid.tolist())

    X_out = X.copy()
    row_keep_for_third_svd = np.ones(X.shape[0], dtype=bool)
    warnings_list: List[str] = []

    for i in range(X.shape[0]):
        fit_valid_bins = [
            b for b in fit_bins.tolist()
            if b not in original_mask_set and b not in invalid_mask_set and np.isfinite(X[i, b])
        ]
        if not fit_valid_bins:
            row_keep_for_third_svd[i] = False
            warnings_list.append(f"row {i}: no valid bins in 0.3--3 MHz to estimate low-frequency plateau")
            continue

        plateau = float(np.median(X[i, fit_valid_bins]))
        if not np.isfinite(plateau):
            row_keep_for_third_svd[i] = False
            warnings_list.append(f"row {i}: non-finite plateau estimate in 0.3--3 MHz")
            continue

        X_out[i, low_fill_bins] = plateau

        for b in transition_bins.tolist():
            if b in original_mask_set or b in invalid_mask_set:
                continue
            f = freqs[b]
            t = (f - fit_hi_mhz) / max(transition_hi_mhz - fit_hi_mhz, 1e-12)
            t = min(max(t, 0.0), 1.0)
            w = 0.5 * (1.0 + np.cos(np.pi * t))
            X_out[i, b] = w * plateau + (1.0 - w) * X[i, b]

    if third_svd_mask_bins.size > 0:
        X_out[:, third_svd_mask_bins] = 0.0

    X_out = np.maximum(X_out, 0.0)

    n_low_filled = int(np.count_nonzero(freqs < preamp_low_cut_mhz))
    warnings_list.append(f"third_stage_low_fill_bins_below_{preamp_low_cut_mhz:.3f}_MHz={n_low_filled}")
    warnings_list.append(f"third_stage_high_invalid_bins_kept_masked={len(high_invalid_bins)}")
    warnings_list.append(f"third_stage_fit_region_MHz={preamp_low_cut_mhz:.3f}--{fit_hi_mhz:.3f}")
    warnings_list.append(f"third_stage_transition_region_MHz={fit_hi_mhz:.3f}--{transition_hi_mhz:.3f}")
    return X_out, third_svd_mask_bins, row_keep_for_third_svd, warnings_list

# ----------------------------
# Core grouped runner
# ----------------------------

def run_one_nmod4_group(
    nmod4: int,
    group_items: List[Tuple[int, str]],
    out_root: str,
    prod_list: Optional[List[int]],
    mean_pdf: Optional[PdfPages],
    eigvec_pdf: Optional[PdfPages],
    preamp_mean_pdf: Optional[PdfPages],
    preamp_eigvec_pdf: Optional[PdfPages],
    third_mean_pdf: Optional[PdfPages],
    third_eigvec_pdf: Optional[PdfPages],
    stacked_eigen_pdf: Optional[PdfPages],
    session_dir_root: str,
    assignment_rows: List[Dict[str, object]],
    gain_cache: GainModelCache,
    preamp_models: Dict[int, PreampNoiseModel],
) -> None:
    group_name = f"nmod4_{nmod4}"
    group_out = os.path.join(out_root, group_name)
    ensure_dir(group_out)

    sessions_info, n_prod, n_bins = load_group_data(group_items, session_dir_root)

    if prod_list is None:
        products = list(range(n_prod))
    else:
        products = [p for p in prod_list if 0 <= p < n_prod]
        if not products:
            raise ValueError(f"No valid products in {prod_list}; files have {n_prod} products")

    total_rows = sum(int(si["n_time"]) for si in sessions_info)

    with open(os.path.join(group_out, "meta.txt"), "w") as f:
        f.write(f"group: {group_name}\n")
        f.write(f"n_sessions: {len(sessions_info)}\n")
        f.write(f"total_rows_stacked: {total_rows}\n")
        f.write(f"data_shape_per_session: (time, product, bin)\n")
        f.write(f"n_products: {n_prod}\n")
        f.write(f"n_bins: {n_bins}\n")
        f.write(f"products_run: {products}\n")
        f.write("\nSessions in group:\n")
        for si in sessions_info:
            gain_list = si["gain_unique"].tolist() if getattr(si["gain_unique"], "size", 0) else []
            notch_list = si["notch_unique"].tolist() if getattr(si["notch_unique"], "size", 0) else []
            f.write(
                f"  campaign={si['campaign_name']}, "
                f"session_{int(si['session_num']):03d}: "
                f"path={si['path']}, "
                f"n_time={si['n_time']}, "
                f"gain_unique={gain_list}, "
                f"notch_unique={notch_list}\n"
            )

    phase_rows = []
    for r in assignment_rows:
        if r["status"] == "used" and r["assigned_phase"] == nmod4:
            phase_rows.append([
                r["campaign_name"],
                r["campaign_num"],
                r["session_num"],
                r["assigned_phase"],
                r["has_notch_zero"],
                r["status"],
                r["reason"],
                r["path"],
                r["shape"],
                r["notch_unique"],
            ])

    write_csv_rows(
        os.path.join(group_out, "phase_assignment.csv"),
        [
            "campaign_name", "campaign_num", "session_num", "assigned_phase",
            "has_notch_zero", "status", "reason", "path", "shape", "notch_unique"
        ],
        phase_rows,
    )

    manual_mask_bins = np.array([0], dtype=int)
    bin_indices = np.arange(n_bins, dtype=float)
    bin_freqs_mhz = CHANNEL_BIN_MHZ * bin_indices

    for p in products:
        prod_out = os.path.join(group_out, product_dir_name(p))
        ensure_dir(prod_out)

        prod_name = product_label(p)

        X_parts: List[np.ndarray] = []
        row_meta: List[List[object]] = []

        global_row = 0
        for si in sessions_info:
            campaign_name = str(si["campaign_name"])
            session_num = int(si["session_num"])
            path = str(si["path"])
            data = np.asarray(si["data"], dtype=np.float64)

            Xi = data[:, p, :]
            X_parts.append(Xi)

            session_base = os.path.basename(path).replace(".h5", "")
            for local_row in range(Xi.shape[0]):
                row_meta.append([global_row, campaign_name, session_num, session_base, local_row])
                global_row += 1

        if not X_parts:
            print(f"[WARN] {group_name} prod {p}: no session data available; skipping")
            continue

        X = np.vstack(X_parts)

        print(f"[Info] {group_name} prod {p}: stacked X shape before cleaning = {X.shape}")
        print(f"[Info] {group_name} prod {p}: finite fraction in X = {np.isfinite(X).mean():.6f}")

        stacked_model_inputs = collect_stacked_model_inputs_for_product(
            sessions_info=sessions_info,
            X_row_count_expected=X.shape[0],
            product=p,
        )

        X, keep_mask = sanitize_matrix_rows(X)
        row_meta = [row_meta[i] for i in range(len(row_meta)) if keep_mask[i]]
        stacked_model_inputs = subset_model_inputs(stacked_model_inputs, keep_mask)

        print(f"[Info] {group_name} prod {p}: stacked X shape after X-cleaning = {X.shape}")

        if X.shape[0] == 0:
            print(f"[WARN] {group_name} prod {p}: all stacked rows were non-finite; skipping")
            continue

        write_csv_rows(
            os.path.join(prod_out, "row_gain_labels.csv"),
            ["sample", "campaign_name", "session_num", "session_name", "row_in_session", "gain_level"],
            [row_meta[i] + [str(stacked_model_inputs["gain_level"][i])] for i in range(len(row_meta))],
        )

        for gain_level in gain_groups_for_product(p):
            gain_mask = np.asarray(stacked_model_inputs["gain_level"] == gain_level)
            if not np.any(gain_mask):
                print(f"[WARN] {group_name} prod {p} gain {gain_level}: no rows; skipping")
                continue

            gain_out = os.path.join(prod_out, f"gain_{gain_level}")
            ensure_dir(gain_out)

            svd1_out = os.path.join(gain_out, "svd_1")
            svd2_out = os.path.join(gain_out, "svd_2")
            svd3_out = os.path.join(gain_out, "svd_3")
            ensure_dir(svd1_out)
            ensure_dir(svd2_out)
            ensure_dir(svd3_out)

            Xg = X[gain_mask]
            row_meta_g = [row_meta[i] for i in range(len(row_meta)) if gain_mask[i]]
            model_inputs_g = subset_model_inputs(stacked_model_inputs, gain_mask)

            print(f"[Info] {group_name} prod {p} gain {gain_level}: stacked X shape after gain split = {Xg.shape}")

            if p <= 3:
                final_mask_bins, mean_spec_raw, median_spec_raw, filter_stat, filter_stat_label = detect_bad_bins_auto(
                    Xg,
                    manual_mask_bins=manual_mask_bins,
                    start_bin=150,
                    smooth_window=13,
                    baseline_half_window=61,
                    absolute_stat_threshold=1.0,
                    iterative_remaining_threshold=0.25,
                    min_run_length=2,
                    mask_pad=0,
                    iterative_pad=0,
                )
                residual_mode = "fractional"
                first_svd_mean_ylabel = "nV/√Hz"
                first_svd_residual_ylabel = "Frac resid"
            else:
                final_mask_bins, mean_spec_raw, median_spec_raw, filter_stat, filter_stat_label = detect_bad_bins_cross(
                    Xg,
                    manual_mask_bins=manual_mask_bins,
                    start_bin=150,
                    smooth_window=13,
                    baseline_half_window=61,
                    absolute_stat_threshold=1.0,
                    iterative_remaining_threshold=0.25,
                    min_run_length=2,
                    mask_pad=0,
                    iterative_pad=0,
                )
                residual_mode = "absolute"
                first_svd_mean_ylabel = "nV/√Hz"
                first_svd_residual_ylabel = "Abs resid"

            write_csv_rows(
                os.path.join(gain_out, "masked_bins_final.csv"),
                ["bin_index"],
                [[int(b)] for b in final_mask_bins],
            )

            write_csv_rows(
                os.path.join(gain_out, "mean_median_filter_stat.csv"),
                ["bin_index", "mean", "median", "filter_stat"],
                [[i, float(mean_spec_raw[i]), float(median_spec_raw[i]), float(filter_stat[i])] for i in range(len(filter_stat))],
            )

            with open(os.path.join(gain_out, "filter_method.txt"), "w") as f:
                f.write(f"product={p}\n")
                f.write(f"product_label={prod_name}\n")
                f.write(f"gain_level={gain_level}\n")
                f.write("conversion_after_filtering=nV_per_sqrt_Hz\n")
                if p <= 3:
                    f.write("filter_mode=autocorrelation_absolute_threshold\n")
                    f.write("residual_mode=fractional\n")
                    f.write("conversion_formula=sqrt(original_spectrum / interpolated_gain)\n")
                    f.write("stat=localized_excess_in_smoothed_abs_mean_minus_median_above_moving_median_baseline\n")
                    f.write("start_bin=150\nsmooth_window=13\nbaseline_half_window=61\nabsolute_stat_threshold=1.0\niterative_remaining_threshold=0.25\nmin_run_length=2\nmask_pad=0\niterative_pad=0\n")
                else:
                    f.write("filter_mode=cross_correlation_absolute_threshold\n")
                    f.write("residual_mode=absolute\n")
                    f.write("conversion_formula=sign(original_spectrum) * sqrt(abs(original_spectrum) / sqrt(gain_a * gain_b))\n")
                    f.write("stat=localized_excess_in_smoothed_abs_mean_minus_median_above_moving_median_baseline\n")
                    f.write("start_bin=150\nsmooth_window=13\nbaseline_half_window=61\nabsolute_stat_threshold=1.0\niterative_remaining_threshold=0.25\nmin_run_length=2\nmask_pad=0\niterative_pad=0\n")


            print(
                f"[Info] {group_name} prod {p} gain {gain_level}: final masked {len(final_mask_bins)} bins "
                f"using {filter_stat_label}; residual mode = {residual_mode}"
            )

            plot_mean_median_with_mask(
                mean_spec_raw,
                median_spec_raw,
                final_mask_bins,
                out_png=os.path.join(svd1_out, "mean_median_with_masked_bins.png"),
                title=short_title(group_name, p, "Mean/Median", gain_level),
                ylabel="Spec units",
                force_zero_bottom=(p <= 3),
            )

            plot_stat_with_mask(
                filter_stat,
                final_mask_bins,
                out_png=os.path.join(svd1_out, "mean_median_filter_stat_with_masked_bins.png"),
                title=short_title(group_name, p, "Bad-bin stat", gain_level),
                ylabel="Statistic",
            )

            Xg_converted, conversion_warnings, row_keep_conv, invalid_conversion_bins = build_nv_sqrt_hz_matrix_for_gain_group(
                X=Xg,
                product=p,
                gain_level=gain_level,
                model_inputs=model_inputs_g,
                bin_freqs_mhz=bin_freqs_mhz,
                gain_cache=gain_cache,
            )

            save_preamp_model_summary(
                os.path.join(svd1_out, "conversion_to_nV_sqrtHz_summary.txt"),
                conversion_warnings,
            )

            write_csv_rows(
                os.path.join(svd1_out, "conversion_row_keep_mask.csv"),
                ["sample", "campaign_name", "session_num", "session_name", "row_in_session", "keep_after_conversion"],
                [row_meta_g[i] + [int(bool(row_keep_conv[i]))] for i in range(len(row_meta_g))],
            )

            first_svd_mask_bins = np.union1d(final_mask_bins, invalid_conversion_bins)

            Xg_converted = Xg_converted[row_keep_conv, :]
            row_meta_conv = [row_meta_g[i] for i in range(len(row_meta_g)) if row_keep_conv[i]]
            model_inputs_conv = subset_model_inputs(model_inputs_g, row_keep_conv)

            if Xg_converted.shape[0] == 0:
                print(f"[WARN] {group_name} prod {p} gain {gain_level}: all rows were excluded during nV/sqrt(Hz) conversion")
                continue

            if first_svd_mask_bins.size > 0:
                Xg_converted[:, first_svd_mask_bins] = 0.0

            mean_spec = np.mean(Xg_converted, axis=0)
            median_spec = np.median(Xg_converted, axis=0)

            plot_mean_median_with_mask(
                mean_spec,
                median_spec,
                first_svd_mask_bins,
                out_png=os.path.join(svd1_out, "converted_mean_median_with_masked_bins.png"),
                title=short_title(group_name, p, "Converted Mean", gain_level),
                ylabel="nV/√Hz",
                force_zero_bottom=(p <= 3),
            )

            if residual_mode == "fractional":
                R, mu = fractional_residuals(Xg_converted, mask_bins=first_svd_mask_bins)
            else:
                R, mu = absolute_residuals(Xg_converted, mask_bins=first_svd_mask_bins)

            R, keep_mask_R = sanitize_matrix_rows(R)
            X_for_original_svd = Xg_converted[keep_mask_R]
            row_meta_original = [row_meta_conv[i] for i in range(len(row_meta_conv)) if keep_mask_R[i]]

            print(f"[Info] {group_name} prod {p} gain {gain_level}: residual matrix shape after R-cleaning = {R.shape}")

            if R.shape[0] == 0:
                print(f"[WARN] {group_name} prod {p} gain {gain_level}: no finite residual rows remained after cleaning; skipping")
                continue

            try:
                pcs, eigvecs, eigenvals = svd_pca(R)
            except Exception as e:
                print(f"[WARN] {group_name} prod {p} gain {gain_level}: SVD failed even after cleaning: {e}")
                continue

            save_svd_products(
                X_for_mean=X_for_original_svd,
                R=R,
                mu=mu,
                pcs=pcs,
                eigvecs=eigvecs,
                eigenvals=eigenvals,
                row_meta=row_meta_original,
                prod_out=gain_out,
                filename_prefix="",
                title_prefix=short_title(group_name, p, "SVD1", gain_level),
                residual_mode=residual_mode,
                mean_pdf=mean_pdf,
                eigvec_pdf=eigvec_pdf,
                omitted_bins=first_svd_mask_bins,
                mean_ylabel=first_svd_mean_ylabel,
                residual_ylabel=first_svd_residual_ylabel,
                force_zero_bottom=(p <= 3),
            )

            if p > 3:
                plot_stacked_eigenvalues_first20(
                    eigenvals_1=eigenvals,
                    eigenvals_2=None,
                    eigenvals_3=None,
                    out_pdf=stacked_eigen_pdf,
                    out_png=os.path.join(gain_out, "stacked_eigenvalues_first20.png"),
                    title=short_title(group_name, p, "Eigvals", gain_level),
                )

                model_inputs_final = subset_model_inputs(model_inputs_conv, keep_mask_R)
                run_final_stage_regression_and_outputs(
                    product=p,
                    gain_level=gain_level,
                    final_stage_name="svd_1",
                    final_stage_out=svd1_out,
                    title_prefix=short_title(group_name, p, "Reg SVD1", gain_level),
                    residual_mode=residual_mode,
                    X_final=X_for_original_svd,
                    mean_vec=mu,
                    eigvecs=eigvecs,
                    pcs=pcs,
                    masked_bins=first_svd_mask_bins,
                    row_meta=row_meta_original,
                    model_inputs=model_inputs_final,
                )

            if p <= 3:
                channel = p

                write_model_input_summary_csv(
                    os.path.join(svd2_out, "preamp_model_row_inputs.csv"),
                    row_meta=row_meta_conv,
                    model_inputs=model_inputs_conv,
                    channel=channel,
                )

                preamp_model_matrix, preamp_warnings, row_keep_preamp_ps, invalid_preamp_freq_bins = build_preamp_noise_matrix_for_auto_product(
                    channel=channel,
                    model_inputs=model_inputs_conv,
                    bin_freqs_mhz=bin_freqs_mhz,
                    preamp_models=preamp_models,
                )

                save_preamp_model_summary(
                    os.path.join(svd2_out, "preamp_model_summary.txt"),
                    preamp_warnings,
                )

                preamp_second_svd_mask_bins = np.union1d(first_svd_mask_bins, invalid_preamp_freq_bins)

                if preamp_second_svd_mask_bins.size > 0:
                    preamp_model_matrix[:, preamp_second_svd_mask_bins] = 0.0

                mean_preamp = np.mean(preamp_model_matrix, axis=0)
                np.save(os.path.join(gain_out, "preamp_modeled_mean.npy"), mean_preamp)

                write_csv_rows(
                    os.path.join(svd2_out, "preamp_modeled_mean.csv"),
                    ["bin_index", "frequency_MHz", "mean_preamp_noise_nV_per_sqrtHz"],
                    [[i, float(bin_freqs_mhz[i]), float(mean_preamp[i])] for i in range(n_bins)],
                )

                write_csv_rows(
                    os.path.join(svd2_out, "preamp_subtracted_masked_bins.csv"),
                    ["bin_index"],
                    [[int(b)] for b in preamp_second_svd_mask_bins],
                )

                write_csv_rows(
                    os.path.join(svd2_out, "preamp_subtracted_row_keep_mask.csv"),
                    ["sample", "campaign_name", "session_num", "session_name", "row_in_session", "keep_for_second_svd"],
                    [row_meta_conv[i] + [int(bool(row_keep_preamp_ps[i]))] for i in range(len(row_meta_conv))],
                )

                plot_mean_with_preamp_overlay(
                    mean_original=mean_spec,
                    mean_preamp=mean_preamp,
                    masked_bins=preamp_second_svd_mask_bins,
                    out_png=os.path.join(svd2_out, "mean_with_preamp_overlay_masked.png"),
                    title=short_title(group_name, p, "Mean vs Preamp", gain_level),
                    ylabel="nV/√Hz",
                    force_zero_bottom=True,
                )

                X_preamp_subtracted_raw = Xg_converted - preamp_model_matrix
                X_preamp_subtracted_raw = X_preamp_subtracted_raw[row_keep_preamp_ps, :]
                row_meta_ps = [row_meta_conv[i] for i in range(len(row_meta_conv)) if row_keep_preamp_ps[i]]

                if X_preamp_subtracted_raw.shape[0] == 0:
                    print(f"[WARN] {group_name} prod {p} gain {gain_level}: all rows were excluded before second SVD")
                    continue

                X_preamp_subtracted = X_preamp_subtracted_raw.copy()
                if preamp_second_svd_mask_bins.size > 0:
                    X_preamp_subtracted[:, preamp_second_svd_mask_bins] = 0.0

                mean_spec_ps = np.mean(X_preamp_subtracted, axis=0)
                median_spec_ps = np.median(X_preamp_subtracted, axis=0)
                plot_mean_median_with_mask(
                    mean_spec_ps,
                    median_spec_ps,
                    preamp_second_svd_mask_bins,
                    out_png=os.path.join(svd2_out, "preamp_subtracted_mean_median_with_original_masked_bins.png"),
                    title=short_title(group_name, p, "Preamp Mean", gain_level),
                    ylabel="nV/√Hz",
                    force_zero_bottom=True,
                )

                R_ps, mu_ps = fractional_residuals(X_preamp_subtracted, mask_bins=preamp_second_svd_mask_bins)
                R_ps, keep_mask_R_ps = sanitize_matrix_rows(R_ps)
                X_ps_for_svd = X_preamp_subtracted[keep_mask_R_ps]
                row_meta_ps_second = [row_meta_ps[i] for i in range(len(row_meta_ps)) if keep_mask_R_ps[i]]

                print(f"[Info] {group_name} prod {p} gain {gain_level}: preamp-subtracted residual matrix shape after cleaning = {R_ps.shape}")

                if R_ps.shape[0] == 0:
                    print(f"[WARN] {group_name} prod {p} gain {gain_level}: no finite preamp-subtracted residual rows remained; skipping second SVD")
                    continue

                try:
                    pcs_ps, eigvecs_ps, eigenvals_ps = svd_pca(R_ps)
                except Exception as e:
                    print(f"[WARN] {group_name} prod {p} gain {gain_level}: preamp-subtracted SVD failed: {e}")
                    continue

                save_svd_products(
                    X_for_mean=X_ps_for_svd,
                    R=R_ps,
                    mu=mu_ps,
                    pcs=pcs_ps,
                    eigvecs=eigvecs_ps,
                    eigenvals=eigenvals_ps,
                    row_meta=row_meta_ps_second,
                    prod_out=svd2_out,
                    filename_prefix="",
                    title_prefix=short_title(group_name, p, "Preamp", gain_level),
                    residual_mode="fractional",
                    mean_pdf=preamp_mean_pdf,
                    eigvec_pdf=preamp_eigvec_pdf,
                    omitted_bins=preamp_second_svd_mask_bins,
                    mean_ylabel="nV/√Hz",
                    residual_ylabel="Frac resid",
                    force_zero_bottom=True,
                )

                X_third_stage, third_svd_mask_bins, row_keep_third_stage, third_stage_warnings = build_lowfreq_repaired_matrix_for_third_svd(
                    X_preamp_subtracted_raw=X_preamp_subtracted_raw,
                    bin_freqs_mhz=bin_freqs_mhz,
                    original_mask_bins=first_svd_mask_bins,
                    invalid_preamp_freq_bins=invalid_preamp_freq_bins,
                    fit_hi_mhz=3.0,
                    transition_hi_mhz=5.0,
                    preamp_low_cut_mhz=0.3,
                )

                save_preamp_model_summary(
                    os.path.join(svd3_out, "third_stage_lowfreq_summary.txt"),
                    third_stage_warnings,
                )

                write_csv_rows(
                    os.path.join(svd3_out, "third_stage_masked_bins.csv"),
                    ["bin_index"],
                    [[int(b)] for b in third_svd_mask_bins],
                )

                write_csv_rows(
                    os.path.join(svd3_out, "third_stage_row_keep_mask.csv"),
                    ["sample", "campaign_name", "session_num", "session_name", "row_in_session", "keep_for_third_svd"],
                    [row_meta_ps[i] + [int(bool(row_keep_third_stage[i]))] for i in range(len(row_meta_ps))],
                )

                mean_spec_third = np.mean(X_third_stage, axis=0)
                median_spec_third = np.median(X_third_stage, axis=0)
                plot_mean_median_with_mask(
                    mean_spec_third,
                    median_spec_third,
                    third_svd_mask_bins,
                    out_png=os.path.join(svd3_out, "third_stage_mean_median_with_mask.png"),
                    title=short_title(group_name, p, "Repaired Mean", gain_level),
                    ylabel="nV/√Hz",
                )

                plot_mean_comparison(
                    mean_a=mean_spec_ps,
                    mean_b=mean_spec_third,
                    masked_bins=third_svd_mask_bins,
                    out_png=os.path.join(svd3_out, "third_stage_vs_preamp_subtracted_mean.png"),
                    title=short_title(group_name, p, "Preamp vs Repair", gain_level),
                    label_a="Preamp-subtracted mean",
                    label_b="Low-frequency repaired mean",
                    ylabel="nV/√Hz",
                )

                X_third_stage = X_third_stage[row_keep_third_stage, :]
                row_meta_third = [row_meta_ps[i] for i in range(len(row_meta_ps)) if row_keep_third_stage[i]]

                if X_third_stage.shape[0] == 0:
                    print(f"[WARN] {group_name} prod {p} gain {gain_level}: all rows were excluded before third SVD")
                    continue

                R_third, mu_third = fractional_residuals(X_third_stage, mask_bins=third_svd_mask_bins)
                R_third, keep_mask_R_third = sanitize_matrix_rows(R_third)
                X_third_for_svd = X_third_stage[keep_mask_R_third]
                row_meta_third = [row_meta_third[i] for i in range(len(row_meta_third)) if keep_mask_R_third[i]]

                print(f"[Info] {group_name} prod {p} gain {gain_level}: third-stage residual matrix shape after cleaning = {R_third.shape}")

                if R_third.shape[0] == 0:
                    print(f"[WARN] {group_name} prod {p} gain {gain_level}: no finite third-stage residual rows remained; skipping third SVD")
                    continue

                try:
                    pcs_third, eigvecs_third, eigenvals_third = svd_pca(R_third)
                except Exception as e:
                    print(f"[WARN] {group_name} prod {p} gain {gain_level}: third-stage SVD failed: {e}")
                    continue

                save_svd_products(
                    X_for_mean=X_third_for_svd,
                    R=R_third,
                    mu=mu_third,
                    pcs=pcs_third,
                    eigvecs=eigvecs_third,
                    eigenvals=eigenvals_third,
                    row_meta=row_meta_third,
                    prod_out=svd3_out,
                    filename_prefix="",
                    title_prefix=short_title(group_name, p, "Repair", gain_level),
                    residual_mode="fractional",
                    mean_pdf=third_mean_pdf,
                    eigvec_pdf=third_eigvec_pdf,
                    omitted_bins=third_svd_mask_bins,
                    mean_ylabel="nV/√Hz",
                    residual_ylabel="Frac resid",
                    force_zero_bottom=True,
                )
                plot_stacked_eigenvalues_first20(
                    eigenvals_1=eigenvals,
                    eigenvals_2=eigenvals_ps,
                    eigenvals_3=eigenvals_third,
                    out_pdf=stacked_eigen_pdf,
                    out_png=os.path.join(gain_out, "stacked_eigenvalues_first20.png"),
                    title=short_title(group_name, p, "Eigvals", gain_level),
                )

                model_inputs_ps = subset_model_inputs(model_inputs_conv, row_keep_preamp_ps)
                model_inputs_third = subset_model_inputs(model_inputs_ps, row_keep_third_stage)
                model_inputs_third = subset_model_inputs(model_inputs_third, keep_mask_R_third)

                run_final_stage_regression_and_outputs(
                    product=p,
                    gain_level=gain_level,
                    final_stage_name="svd_3",
                    final_stage_out=svd3_out,
                    title_prefix=short_title(group_name, p, "Reg Repair", gain_level),
                    residual_mode="fractional",
                    X_final=X_third_for_svd,
                    mean_vec=mu_third,
                    eigvecs=eigvecs_third,
                    pcs=pcs_third,
                    masked_bins=third_svd_mask_bins,
                    row_meta=row_meta_third,
                    model_inputs=model_inputs_third,
                )



# ----------------------------
# Regression / reconstruction helpers for final-stage SVD products
# ----------------------------

def get_regression_telemetry_matrix_for_product(
    product: int,
    model_inputs: Dict[str, np.ndarray],
) -> Tuple[np.ndarray, List[str]]:
    """
    Build the telemetry design columns used for the PC regressions.
    We use the same core terms as the gain model, but collapse the ADC choice
    into a generic TADC column so mixed cross-products can still be handled.

    Autos:
      ch0/ch1 -> TADC = SPE_ADC0_T
      ch2/ch3 -> TADC = SPE_ADC1_T

    Cross products:
      if both channels live on the same ADC side, use that ADC;
      otherwise use the average of SPE_ADC0_T and SPE_ADC1_T.
    """
    tele_cols = ["THERM_FPGA", "TADC", "SPE_1VAD8_V", "VMON_1V2D", "SPE_1VAD8_C"]

    adc0 = np.asarray(model_inputs["SPE_ADC0_T"], dtype=float)
    adc1 = np.asarray(model_inputs["SPE_ADC1_T"], dtype=float)

    if product <= 3:
        tadc = adc0 if product in (0, 1) else adc1
    else:
        ch_a, ch_b = CROSS_PRODUCT_CHANNELS[product]
        if ch_a in (0, 1) and ch_b in (0, 1):
            tadc = adc0
        elif ch_a in (2, 3) and ch_b in (2, 3):
            tadc = adc1
        else:
            tadc = 0.5 * (adc0 + adc1)

    Xtele = np.column_stack([
        np.asarray(model_inputs["THERM_FPGA"], dtype=float),
        np.asarray(tadc, dtype=float),
        np.asarray(model_inputs["SPE_1VAD8_V"], dtype=float),
        np.asarray(model_inputs["VMON_1V2D"], dtype=float),
        np.asarray(model_inputs["SPE_1VAD8_C"], dtype=float),
    ])
    return Xtele, tele_cols


def reconstruct_spectra_from_pc_scores(
    pred_pcs_12: np.ndarray,
    mean_vec: np.ndarray,
    eigvecs: np.ndarray,
    residual_mode: str,
) -> np.ndarray:
    pred_pcs_12 = np.asarray(pred_pcs_12, dtype=float)
    mean_vec = np.asarray(mean_vec, dtype=float).reshape(-1)
    eigvecs = np.asarray(eigvecs, dtype=float)

    k = min(2, eigvecs.shape[1], pred_pcs_12.shape[1])
    if k == 0:
        return np.tile(mean_vec.reshape(1, -1), (pred_pcs_12.shape[0], 1))

    resid_hat = pred_pcs_12[:, :k] @ eigvecs[:, :k].T

    if residual_mode == "fractional":
        return mean_vec.reshape(1, -1) * (1.0 + resid_hat)
    if residual_mode == "absolute":
        return mean_vec.reshape(1, -1) + resid_hat

    raise ValueError(f"Unsupported residual_mode: {residual_mode}")


def compute_rowwise_nrms(
    actual: np.ndarray,
    predicted: np.ndarray,
    masked_bins: Optional[np.ndarray] = None,
) -> np.ndarray:
    actual = np.asarray(actual, dtype=float)
    predicted = np.asarray(predicted, dtype=float)

    if actual.shape != predicted.shape:
        raise ValueError(f"Shape mismatch in NRMS computation: {actual.shape} vs {predicted.shape}")

    n_rows, n_bins = actual.shape
    valid_bins = np.ones(n_bins, dtype=bool)
    if masked_bins is not None and np.size(masked_bins) > 0:
        masked_bins = np.asarray(masked_bins, dtype=int)
        good = (masked_bins >= 0) & (masked_bins < n_bins)
        valid_bins[masked_bins[good]] = False

    out = np.full(n_rows, np.nan, dtype=float)
    for i in range(n_rows):
        row_ok = valid_bins & np.isfinite(actual[i]) & np.isfinite(predicted[i])
        if not np.any(row_ok):
            continue
        denom = float(np.mean(actual[i, row_ok]))
        if not np.isfinite(denom) or abs(denom) <= 0.0:
            continue
        out[i] = float(np.sqrt(np.mean((actual[i, row_ok] - predicted[i, row_ok]) ** 2)) / denom)
    return out


def compute_per_frequency_rms(
    actual: np.ndarray,
    predicted: np.ndarray,
    masked_bins: Optional[np.ndarray] = None,
) -> np.ndarray:
    actual = np.asarray(actual, dtype=float)
    predicted = np.asarray(predicted, dtype=float)

    rms = np.sqrt(np.nanmean((actual - predicted) ** 2, axis=0))
    if masked_bins is not None and np.size(masked_bins) > 0:
        masked_bins = np.asarray(masked_bins, dtype=int)
        good = (masked_bins >= 0) & (masked_bins < rms.size)
        rms[masked_bins[good]] = np.nan
    return rms


def plot_pc_actual_vs_predicted(
    y_true: np.ndarray,
    y_pred: np.ndarray,
    out_png: str,
    title: str,
    xlabel: str,
    ylabel: str,
) -> None:
    y_true = np.asarray(y_true, dtype=float)
    y_pred = np.asarray(y_pred, dtype=float)

    fig = plt.figure(figsize=(6, 6))
    plt.scatter(y_true, y_pred, alpha=0.65)
    both = np.concatenate([y_true[np.isfinite(y_true)], y_pred[np.isfinite(y_pred)]])
    if both.size > 0:
        lo = float(np.min(both))
        hi = float(np.max(both))
        if np.isfinite(lo) and np.isfinite(hi) and hi > lo:
            plt.plot([lo, hi], [lo, hi], "k--", linewidth=1.0)
    plt.xlabel(xlabel)
    plt.ylabel(ylabel)
    plt.title(title)
    plt.grid(True, alpha=0.3)
    plt.tight_layout()
    plt.savefig(out_png)
    plt.close(fig)


def plot_nrms_distribution_with_truncation(
    trunc_nrms: np.ndarray,
    pred_nrms: np.ndarray,
    out_png: str,
    title: str,
) -> None:
    trunc_nrms = np.asarray(trunc_nrms, dtype=float)
    pred_nrms = np.asarray(pred_nrms, dtype=float)

    good = np.isfinite(trunc_nrms) & np.isfinite(pred_nrms)
    trunc_nrms = trunc_nrms[good]
    pred_nrms = pred_nrms[good]

    if trunc_nrms.size == 0:
        return

    extra = np.maximum(pred_nrms - trunc_nrms, 0.0)
    order = np.argsort(pred_nrms)

    fig = plt.figure(figsize=(10, 5.5))
    x = np.arange(order.size)
    plt.bar(x, trunc_nrms[order], label="Truncation error", alpha=0.85)
    plt.bar(x, extra[order], bottom=trunc_nrms[order], label="Additional regression error", alpha=0.85)
    plt.xlabel("Rows")
    plt.ylabel("nRMS")
    plt.title(title)
    plt.grid(True, axis="y", alpha=0.3)
    plt.legend()
    plt.ylim(bottom=0)
    plt.tight_layout()
    plt.savefig(out_png)
    plt.close(fig)


def plot_nrms_hist_decomposed(
    trunc_nrms: np.ndarray,
    pred_nrms: np.ndarray,
    out_png: str,
    title: str,
) -> None:
    trunc_nrms = np.asarray(trunc_nrms, dtype=float)
    pred_nrms = np.asarray(pred_nrms, dtype=float)

    good = np.isfinite(trunc_nrms) & np.isfinite(pred_nrms)
    if not np.any(good):
        return

    t = trunc_nrms[good]
    p = pred_nrms[good]
    extra = np.maximum(p - t, 0.0)

    maxv = float(np.max(p)) if p.size > 0 else 0.0
    bins = np.linspace(0.0, max(maxv * 1.05, 1e-6), 25)
    bin_centers = 0.5 * (bins[:-1] + bins[1:])
    bin_widths = np.diff(bins)

    trunc_part = np.zeros(len(bin_centers), dtype=float)
    extra_part = np.zeros(len(bin_centers), dtype=float)

    idx = np.digitize(p, bins) - 1
    idx = np.clip(idx, 0, len(bin_centers) - 1)

    for k in range(len(bin_centers)):
        m = idx == k
        if not np.any(m):
            continue

        p_bin = p[m]
        t_bin = t[m]
        e_bin = extra[m]

        denom = np.where(p_bin > 0.0, p_bin, 1.0)
        trunc_frac = np.clip(t_bin / denom, 0.0, 1.0)
        extra_frac = np.clip(e_bin / denom, 0.0, 1.0)

        trunc_part[k] = np.sum(trunc_frac)
        extra_part[k] = np.sum(extra_frac)

    fig = plt.figure(figsize=(8.5, 5.2))
    plt.bar(
        bin_centers,
        trunc_part,
        width=bin_widths,
        align="center",
        label="Truncation part",
    )
    plt.bar(
        bin_centers,
        extra_part,
        width=bin_widths,
        align="center",
        bottom=trunc_part,
        label="Regression part",
    )
    plt.xlabel("nRMS")
    plt.ylabel("Count")
    plt.title(title)
    plt.grid(True, axis="y", alpha=0.3)
    plt.legend()
    plt.tight_layout()
    plt.savefig(out_png)
    plt.close(fig)



def plot_nrms_hist_trunc_only(
    trunc_nrms: np.ndarray,
    out_png: str,
    title: str,
) -> None:
    t = np.asarray(trunc_nrms, dtype=float)
    t = t[np.isfinite(t)]
    if t.size == 0:
        return

    maxv = float(np.max(t))
    bins = np.linspace(0.0, max(maxv * 1.05, 1e-6), 25)

    fig = plt.figure(figsize=(8.5, 5.2))
    plt.hist(t, bins=bins, label="Truncation")
    plt.xlabel("nRMS")
    plt.ylabel("Count")
    plt.title(title)
    plt.grid(True, axis="y", alpha=0.3)
    plt.legend()
    plt.tight_layout()
    plt.savefig(out_png)
    plt.close(fig)


def plot_residual_mean_and_per_frequency_rms(
    actual: np.ndarray,
    predicted: np.ndarray,
    masked_bins: np.ndarray,
    out_png_mean: str,
    out_png_rms: str,
    title_prefix: str,
    ylabel: str = "nV/sqrt(Hz)",
) -> None:
    x = np.arange(actual.shape[1]) * CHANNEL_BIN_MHZ
    residual = np.asarray(actual, dtype=float) - np.asarray(predicted, dtype=float)
    mean_resid = np.nanmean(residual, axis=0)
    rms_f = compute_per_frequency_rms(actual, predicted, masked_bins=masked_bins)

    if masked_bins is not None and np.size(masked_bins) > 0:
        masked_bins = np.asarray(masked_bins, dtype=int)
        good = (masked_bins >= 0) & (masked_bins < mean_resid.size)
        mean_resid[masked_bins[good]] = np.nan

    fig = plt.figure(figsize=(10, 5.2))
    plt.plot(x, mean_resid, linewidth=1.3)
    plt.xlabel("Frequency (MHz)")
    plt.ylabel(f"Mean resid [{ylabel}]")
    plt.title(short_stage_title(title_prefix, "Mean resid"))
    plt.grid(True, alpha=0.3)
    plt.tight_layout()
    plt.savefig(out_png_mean)
    plt.close(fig)

    fig = plt.figure(figsize=(10, 5.2))
    plt.plot(x, rms_f, linewidth=1.3)
    plt.xlabel("Frequency (MHz)")
    plt.ylabel(f"RMS [{ylabel}]")
    plt.title(short_stage_title(title_prefix, "RMS resid"))
    plt.grid(True, alpha=0.3)
    plt.ylim(bottom=0)
    plt.tight_layout()
    plt.savefig(out_png_rms)
    plt.close(fig)


def plot_example_row_comparisons(
    actual: np.ndarray,
    trunc_pred: np.ndarray,
    pred: np.ndarray,
    nrms_pred: np.ndarray,
    masked_bins: np.ndarray,
    out_png: str,
    title: str,
    ylabel: str = "nV/sqrt(Hz)",
) -> None:
    actual = np.asarray(actual, dtype=float)
    trunc_pred = np.asarray(trunc_pred, dtype=float)
    pred = np.asarray(pred, dtype=float)
    nrms_pred = np.asarray(nrms_pred, dtype=float)

    good = np.where(np.isfinite(nrms_pred))[0]
    if good.size == 0:
        return

    order = good[np.argsort(nrms_pred[good])]
    picks = [int(order[0]), int(order[len(order)//2]), int(order[-1])]
    labels = ["Best", "Median", "Worst"]

    x = np.arange(actual.shape[1]) * CHANNEL_BIN_MHZ
    mask_plot = np.ones(actual.shape[1], dtype=bool)
    if masked_bins is not None and np.size(masked_bins) > 0:
        masked_bins = np.asarray(masked_bins, dtype=int)
        valid = (masked_bins >= 0) & (masked_bins < actual.shape[1])
        mask_plot[masked_bins[valid]] = False

    fig, axes = plt.subplots(3, 1, figsize=(10, 11), sharex=True)
    for ax, idx, label in zip(axes, picks, labels):
        y_act = actual[idx].copy()
        y_tr = trunc_pred[idx].copy()
        y_pr = pred[idx].copy()
        y_act[~mask_plot] = np.nan
        y_tr[~mask_plot] = np.nan
        y_pr[~mask_plot] = np.nan

        ax.plot(x, y_act, label="Actual", linewidth=1.5)
        ax.plot(x, y_tr, label="Actual PC1+PC2", linestyle=":")
        ax.plot(x, y_pr, label="Predicted PC1+PC2", linestyle="--")
        ax.set_ylabel(ylabel)
        ax.set_title(f"{label} • i={idx} • nRMS={nrms_pred[idx]:.4g}")
        ax.grid(True, alpha=0.3)
        ax.legend()

    axes[-1].set_xlabel("Frequency (MHz)")
    fig.suptitle(title, y=0.995)
    fig.tight_layout(rect=(0, 0, 1, 0.98))
    fig.savefig(out_png)
    plt.close(fig)


def fit_pc_regression_models(
    product: int,
    pcs: np.ndarray,
    model_inputs: Dict[str, np.ndarray],
    jackknife_min_rows: int = JACKKNIFE_MIN_ROWS,
    threshold_ratio: float = THRESHOLD_RATIO,
) -> Tuple[Dict[str, Dict[str, np.ndarray]], List[Dict[str, object]], np.ndarray, List[str]]:
    """
    Fit PC1 and PC2 using linear and quadratic regressions.
    Returns:
      predictions[pc_name][model_name] = predicted PC vector
      alpha_records = exported coefficient / stderr metadata
      valid_rows = rows with finite telemetry and finite PC1/PC2
      tele_cols = labels used in the design matrix
    """
    Xtele, tele_cols = get_regression_telemetry_matrix_for_product(product, model_inputs)
    pcs = np.asarray(pcs, dtype=float)
    valid_rows = np.all(np.isfinite(Xtele), axis=1) & np.all(np.isfinite(pcs[:, :2]), axis=1)

    Xtele = Xtele[valid_rows]
    pcs_valid = pcs[valid_rows, :]

    predictions = {pc: {} for pc in ("PC1", "PC2")}
    alpha_records: List[Dict[str, object]] = []

    for ipc, pc_name in enumerate(("PC1", "PC2")):
        if ipc >= pcs_valid.shape[1]:
            continue
        y = pcs_valid[:, ipc]
        n = y.size
        if n == 0:
            continue

        for order in (1, 2):
            model_name = "linear" if order == 1 else "quadratic"
            Z, labels = build_feature_matrix(Xtele, tele_cols, order=order)

            use_jackknife = bool(n < jackknife_min_rows)
            if use_jackknife and n >= 3:
                alpha_full = np.linalg.lstsq(Z, y, rcond=None)[0]
                jack = []
                for i in range(n):
                    keep = np.ones(n, dtype=bool)
                    keep[i] = False
                    jack.append(np.linalg.lstsq(Z[keep], y[keep], rcond=None)[0])
                jack = np.vstack(jack)
                stderr = np.sqrt(np.var(jack, axis=0, ddof=1) / n)

                keep_idx = []
                for i, (lbl, a, se) in enumerate(zip(labels, alpha_full, stderr)):
                    ratio = (abs(a / se) if se != 0 else np.inf)
                    alpha_records.append({
                        "pc": pc_name,
                        "model": model_name,
                        "term": lbl,
                        "alpha": float(a),
                        "stderr": float(se),
                        "ratio": float(ratio),
                        "used_jackknife": 1,
                        "kept": int(bool(ratio > threshold_ratio)),
                        "n_rows": int(n),
                    })
                    if ratio > threshold_ratio:
                        keep_idx.append(i)

                if keep_idx:
                    Zf = Z[:, keep_idx]
                    af = np.linalg.lstsq(Zf, y, rcond=None)[0]
                    y_pred = Zf @ af
                else:
                    y_pred = np.full(n, np.mean(y), dtype=float)
            else:
                alpha_full = np.linalg.lstsq(Z, y, rcond=None)[0]
                y_pred = Z @ alpha_full
                for lbl, a in zip(labels, alpha_full):
                    alpha_records.append({
                        "pc": pc_name,
                        "model": model_name,
                        "term": lbl,
                        "alpha": float(a),
                        "stderr": np.nan,
                        "ratio": np.nan,
                        "used_jackknife": 0,
                        "kept": 1,
                        "n_rows": int(n),
                    })

            predictions[pc_name][model_name] = y_pred

    return predictions, alpha_records, valid_rows, tele_cols


def run_final_stage_regression_and_outputs(
    product: int,
    gain_level: str,
    final_stage_name: str,
    final_stage_out: str,
    title_prefix: str,
    residual_mode: str,
    X_final: np.ndarray,
    mean_vec: np.ndarray,
    eigvecs: np.ndarray,
    pcs: np.ndarray,
    masked_bins: np.ndarray,
    row_meta: List[List[object]],
    model_inputs: Dict[str, np.ndarray],
) -> None:
    regression_out = os.path.join(final_stage_out, "regression")
    ensure_dir(regression_out)

    predictions, alpha_records, valid_rows, tele_cols = fit_pc_regression_models(
        product=product,
        pcs=pcs,
        model_inputs=model_inputs,
    )

    write_csv_rows(
        os.path.join(regression_out, "telemetry_columns_used.csv"),
        ["column_name"],
        [[c] for c in tele_cols],
    )

    if alpha_records:
        alpha_df = pd.DataFrame(alpha_records)
        alpha_df.to_csv(os.path.join(regression_out, "pc_regression_alphas.csv"), index=False)

    X_final_valid = X_final[valid_rows]
    pcs_valid = pcs[valid_rows, :]
    row_meta_valid = [row_meta[i] for i in range(len(row_meta)) if valid_rows[i]]

    model_inputs_valid = subset_model_inputs(model_inputs, valid_rows)
    Xtele_valid, _ = get_regression_telemetry_matrix_for_product(product, model_inputs_valid)
    write_csv_rows(
        os.path.join(regression_out, "regression_rows.csv"),
        ["sample", "campaign_name", "session_num", "session_name", "row_in_session", "THERM_FPGA", "TADC", "SPE_1VAD8_V", "VMON_1V2D", "SPE_1VAD8_C"],
        [row_meta_valid[i] + Xtele_valid[i].tolist() for i in range(len(row_meta_valid))],
    )

    actual_pcs_12 = np.zeros((pcs_valid.shape[0], 2), dtype=float)
    actual_pcs_12[:, 0] = pcs_valid[:, 0] if pcs_valid.shape[1] > 0 else 0.0
    actual_pcs_12[:, 1] = pcs_valid[:, 1] if pcs_valid.shape[1] > 1 else 0.0
    actual_pc12_recon = reconstruct_spectra_from_pc_scores(actual_pcs_12, mean_vec, eigvecs, residual_mode=residual_mode)

    trunc_nrms = compute_rowwise_nrms(X_final_valid, actual_pc12_recon, masked_bins=masked_bins)
    pd.DataFrame({
        "sample": [rm[0] for rm in row_meta_valid],
        "campaign_name": [rm[1] for rm in row_meta_valid],
        "session_num": [rm[2] for rm in row_meta_valid],
        "session_name": [rm[3] for rm in row_meta_valid],
        "row_in_session": [rm[4] for rm in row_meta_valid],
        "nrms_truncation_pc12": trunc_nrms,
    }).to_csv(os.path.join(regression_out, "truncation_nrms_rows.csv"), index=False)

    np.save(os.path.join(regression_out, "actual_pc12_reconstruction.npy"), actual_pc12_recon)

    mean_actual = np.nanmean(X_final_valid, axis=0)
    mean_trunc = np.nanmean(actual_pc12_recon, axis=0)
    plot_mean_comparison(
        mean_a=mean_actual,
        mean_b=mean_trunc,
        masked_bins=masked_bins,
        out_png=os.path.join(regression_out, "mean_actual_vs_actual_pc12.png"),
        title=short_stage_title(title_prefix, "Actual vs PC12"),
        label_a="Actual mean",
        label_b="Actual PC1+PC2 mean",
        ylabel="nV/√Hz",
    )

    metrics_rows = []
    for model_key, model_label in (("linear", "Linear"), ("quadratic", "Quadratic")):
        if model_key not in predictions.get("PC1", {}) and model_key not in predictions.get("PC2", {}):
            continue

        pred_pcs_12 = np.column_stack([
            predictions.get("PC1", {}).get(model_key, np.zeros(X_final_valid.shape[0], dtype=float)),
            predictions.get("PC2", {}).get(model_key, np.zeros(X_final_valid.shape[0], dtype=float)),
        ])
        pred_spectra = reconstruct_spectra_from_pc_scores(pred_pcs_12, mean_vec, eigvecs, residual_mode=residual_mode)
        pred_nrms = compute_rowwise_nrms(X_final_valid, pred_spectra, masked_bins=masked_bins)

        np.save(os.path.join(regression_out, f"{model_key}_predicted_pc12_spectra.npy"), pred_spectra)
        write_csv_rows(
            os.path.join(regression_out, f"{model_key}_predicted_pc12_pcs.csv"),
            ["sample", "campaign_name", "session_num", "session_name", "row_in_session", "PC1_pred", "PC2_pred"],
            [row_meta_valid[i] + [float(pred_pcs_12[i, 0]), float(pred_pcs_12[i, 1])] for i in range(len(row_meta_valid))],
        )

        pd.DataFrame({
            "sample": [rm[0] for rm in row_meta_valid],
            "campaign_name": [rm[1] for rm in row_meta_valid],
            "session_num": [rm[2] for rm in row_meta_valid],
            "session_name": [rm[3] for rm in row_meta_valid],
            "row_in_session": [rm[4] for rm in row_meta_valid],
            "nrms_truncation_pc12": trunc_nrms,
            f"nrms_total_{model_key}_pc12": pred_nrms,
            f"nrms_additional_{model_key}": pred_nrms - trunc_nrms,
        }).to_csv(os.path.join(regression_out, f"{model_key}_nrms_rows.csv"), index=False)

        plot_pc_actual_vs_predicted(
            pcs_valid[:, 0] if pcs_valid.shape[1] > 0 else np.zeros(X_final_valid.shape[0]),
            pred_pcs_12[:, 0],
            out_png=os.path.join(regression_out, f"{model_key}_pc1_actual_vs_predicted.png"),
            title=short_stage_title(title_prefix, f"{model_label} PC1"),
            xlabel="Actual PC1",
            ylabel="Pred PC1",
        )
        plot_pc_actual_vs_predicted(
            pcs_valid[:, 1] if pcs_valid.shape[1] > 1 else np.zeros(X_final_valid.shape[0]),
            pred_pcs_12[:, 1],
            out_png=os.path.join(regression_out, f"{model_key}_pc2_actual_vs_predicted.png"),
            title=short_stage_title(title_prefix, f"{model_label} PC2"),
            xlabel="Actual PC2",
            ylabel="Pred PC2",
        )

        plot_nrms_distribution_with_truncation(
            trunc_nrms=trunc_nrms,
            pred_nrms=pred_nrms,
            out_png=os.path.join(regression_out, f"{model_key}_nrms_distribution_stacked.png"),
            title=short_stage_title(title_prefix, f"{model_label} nRMS"),
        )
        plot_nrms_hist_decomposed(
            trunc_nrms=trunc_nrms,
            pred_nrms=pred_nrms,
            out_png=os.path.join(regression_out, f"{model_key}_nrms_hist_decomposed.png"),
            title=short_stage_title(title_prefix, f"{model_label} Hist"),
        )
        plot_nrms_hist_trunc_only(
            trunc_nrms=trunc_nrms,
            out_png=os.path.join(regression_out, f"{model_key}_nrms_hist_trunc.png"),
            title=short_stage_title(title_prefix, f"{model_label} Trunc hist"),
        )

        mean_pred = np.nanmean(pred_spectra, axis=0)
        plot_mean_comparison(
            mean_a=mean_actual,
            mean_b=mean_pred,
            masked_bins=masked_bins,
            out_png=os.path.join(regression_out, f"{model_key}_mean_actual_vs_predicted.png"),
            title=short_stage_title(title_prefix, f"{model_label} Mean"),
            label_a="Actual mean",
            label_b=f"{model_label} predicted mean",
            ylabel="nV/√Hz",
        )

        plot_residual_mean_and_per_frequency_rms(
            actual=X_final_valid,
            predicted=pred_spectra,
            masked_bins=masked_bins,
            out_png_mean=os.path.join(regression_out, f"{model_key}_mean_residual.png"),
            out_png_rms=os.path.join(regression_out, f"{model_key}_per_frequency_rms.png"),
            title_prefix=short_stage_title(title_prefix, model_label),
            ylabel="nV/√Hz",
        )

        plot_example_row_comparisons(
            actual=X_final_valid,
            trunc_pred=actual_pc12_recon,
            pred=pred_spectra,
            nrms_pred=pred_nrms,
            masked_bins=masked_bins,
            out_png=os.path.join(regression_out, f"{model_key}_example_rows.png"),
            title=short_stage_title(title_prefix, f"{model_label} Rows"),
            ylabel="nV/√Hz",
        )

        metrics_rows.append({
            "product": int(product),
            "gain_level": str(gain_level),
            "final_stage": str(final_stage_name),
            "model": model_key,
            "n_rows": int(X_final_valid.shape[0]),
            "median_nrms_truncation": float(np.nanmedian(trunc_nrms)),
            "median_nrms_total": float(np.nanmedian(pred_nrms)),
            "median_nrms_additional": float(np.nanmedian(pred_nrms - trunc_nrms)),
            "global_nrms_truncation": float(np.sqrt(np.nanmean((X_final_valid - actual_pc12_recon) ** 2)) / np.nanmean(X_final_valid)),
            "global_nrms_total": float(np.sqrt(np.nanmean((X_final_valid - pred_spectra) ** 2)) / np.nanmean(X_final_valid)),
        })

    if metrics_rows:
        pd.DataFrame(metrics_rows).to_csv(os.path.join(regression_out, "metrics_summary.csv"), index=False)


# ----------------------------
# CLI main
# ----------------------------

def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument(
        "--session-dir",
        default=os.path.expanduser("~/gain_model/data/h5/Payload_CPTs/Payload_CPTs_Science_h5"),
        help="Root directory containing science_<x> campaign subdirectories.",
    )
    ap.add_argument(
        "--pattern",
        default="session_science_*.h5",
        help="Recursive glob pattern for session HDF5 files beneath --session-dir.",
    )
    ap.add_argument(
        "--out-root",
        default=os.path.expanduser("~/gain_model/outputs/noise_svd_nmod4"),
        help="Root output directory (per-nmod4 subdirs created here).",
    )
    ap.add_argument("--start", type=int, default=0)
    ap.add_argument("--end", type=int, default=999)
    ap.add_argument(
        "--products",
        default="all",
        help="Comma-separated product indices to run or 'all'. Default: all",
    )
    ap.add_argument(
        "--gain-model-root",
        default=DEFAULT_GAIN_MODEL_ROOT,
        help="Root directory of exported corrected gain PCA model products.",
    )
    ap.add_argument(
        "--preamp-xlsx-ch0",
        default=DEFAULT_PREAMP_XLSX[0],
        help="Preamp noise xlsx for channel 0.",
    )
    ap.add_argument(
        "--preamp-xlsx-ch1",
        default=DEFAULT_PREAMP_XLSX[1],
        help="Preamp noise xlsx for channel 1.",
    )
    ap.add_argument(
        "--preamp-xlsx-ch2",
        default=DEFAULT_PREAMP_XLSX[2],
        help="Preamp noise xlsx for channel 2.",
    )
    ap.add_argument(
        "--preamp-xlsx-ch3",
        default=DEFAULT_PREAMP_XLSX[3],
        help="Preamp noise xlsx for channel 3.",
    )
    args = ap.parse_args()

    session_dir_root = os.path.expanduser(args.session_dir)

    files = discover_session_files(session_dir_root, args.pattern)
    if not files:
        raise SystemExit(f"No files matched recursively beneath: {session_dir_root}")

    kept: List[str] = []
    for p in files:
        try:
            n = extract_session_number(p)
        except ValueError:
            continue
        if args.start <= n <= args.end:
            kept.append(p)

    if not kept:
        raise SystemExit(f"Found files, but none in range {args.start:03d}..{args.end:03d}")

    out_root = os.path.expanduser(args.out_root)
    ensure_dir(out_root)

    if args.products.strip().lower() == "all":
        prod_list = None
    else:
        prod_list = [int(x.strip()) for x in args.products.split(",") if x.strip()]

    gain_cache = GainModelCache(args.gain_model_root)

    preamp_models = {
        0: PreampNoiseModel(args.preamp_xlsx_ch0),
        1: PreampNoiseModel(args.preamp_xlsx_ch1),
        2: PreampNoiseModel(args.preamp_xlsx_ch2),
        3: PreampNoiseModel(args.preamp_xlsx_ch3),
    }

    groups, assignment_rows = assign_phases_by_notch(kept, session_dir_root)

    global_phase_rows = []
    for r in assignment_rows:
        global_phase_rows.append([
            r["campaign_name"],
            r["campaign_num"],
            r["session_num"],
            r["assigned_phase"],
            r["has_notch_zero"],
            r["status"],
            r["reason"],
            r["path"],
            r["shape"],
            r["notch_unique"],
        ])

    write_csv_rows(
        os.path.join(out_root, "phase_assignment_all.csv"),
        [
            "campaign_name", "campaign_num", "session_num", "assigned_phase",
            "has_notch_zero", "status", "reason", "path", "shape", "notch_unique"
        ],
        global_phase_rows,
    )

    print(f"[Info] Found {len(kept)} session files in range {args.start:03d}..{args.end:03d}")
    print(f"[Info] Input root:  {session_dir_root}")
    print(f"[Info] Outputs:     {out_root}")
    print(f"[Info] Products:    {'all' if prod_list is None else prod_list}")
    print("[Info] notch-defined groups retained for analysis:")
    for g in [1, 2, 3]:
        print(f"    nmod4={g}: {len(groups.get(g, []))} files")

    mean_pdf_path = os.path.join(out_root, "mean_spectra_all_groups.pdf")
    preamp_mean_pdf_path = os.path.join(out_root, "preamp_subtracted_mean_spectra_all_groups.pdf")
    third_mean_pdf_path = os.path.join(out_root, "third_stage_mean_spectra_all_groups.pdf")

    eigvec_pdfs = {}
    preamp_eigvec_pdfs = {}
    third_eigvec_pdfs = {}
    stacked_eigen_pdfs = {}
    for g in [1, 2, 3]:
        eigvec_pdfs[g] = PdfPages(
            os.path.join(out_root, f"first5_eigenvectors_nmod4_{g}.pdf")
        )
        preamp_eigvec_pdfs[g] = PdfPages(
            os.path.join(out_root, f"preamp_subtracted_first5_eigenvectors_nmod4_{g}.pdf")
        )
        third_eigvec_pdfs[g] = PdfPages(
            os.path.join(out_root, f"third_stage_first5_eigenvectors_nmod4_{g}.pdf")
        )
        stacked_eigen_pdfs[g] = PdfPages(
            os.path.join(out_root, f"stacked_eigenvalues_first20_nmod4_{g}.pdf")
        )

    try:
        with PdfPages(mean_pdf_path) as mean_pdf, PdfPages(preamp_mean_pdf_path) as preamp_mean_pdf, PdfPages(third_mean_pdf_path) as third_mean_pdf:
            for g in [1, 2, 3]:
                if g not in groups:
                    continue
                try:
                    eigvec_pdf = eigvec_pdfs.get(g, None)
                    preamp_eigvec_pdf = preamp_eigvec_pdfs.get(g, None)
                    third_eigvec_pdf = third_eigvec_pdfs.get(g, None)
                    stacked_eigen_pdf = stacked_eigen_pdfs.get(g, None)

                    run_one_nmod4_group(
                        nmod4=g,
                        group_items=groups[g],
                        out_root=out_root,
                        prod_list=prod_list,
                        mean_pdf=mean_pdf,
                        eigvec_pdf=eigvec_pdf,
                        preamp_mean_pdf=preamp_mean_pdf,
                        preamp_eigvec_pdf=preamp_eigvec_pdf,
                        third_mean_pdf=third_mean_pdf,
                        third_eigvec_pdf=third_eigvec_pdf,
                        stacked_eigen_pdf=stacked_eigen_pdf,
                        session_dir_root=session_dir_root,
                        assignment_rows=assignment_rows,
                        gain_cache=gain_cache,
                        preamp_models=preamp_models,
                    )
                except Exception as e:
                    print(f"[FAIL] nmod4={g}: {e}")
    finally:
        for pdf in eigvec_pdfs.values():
            pdf.close()
        for pdf in preamp_eigvec_pdfs.values():
            pdf.close()
        for pdf in third_eigvec_pdfs.values():
            pdf.close()
        for pdf in stacked_eigen_pdfs.values():
            pdf.close()

    print(f"[Done] Wrote combined mean-spectra PDF: {mean_pdf_path}")
    print(f"[Done] Wrote combined preamp-subtracted mean-spectra PDF: {preamp_mean_pdf_path}")
    print(f"[Done] Wrote combined third-stage mean-spectra PDF: {third_mean_pdf_path}")
    for g in [1, 2, 3]:
        if g in groups:
            print(
                f"[Done] Wrote first-5-eigenvectors PDF: "
                f"{os.path.join(out_root, f'first5_eigenvectors_nmod4_{g}.pdf')}"
            )
            print(
                f"[Done] Wrote preamp-subtracted first-5-eigenvectors PDF: "
                f"{os.path.join(out_root, f'preamp_subtracted_first5_eigenvectors_nmod4_{g}.pdf')}"
            )
            print(
                f"[Done] Wrote third-stage first-5-eigenvectors PDF: "
                f"{os.path.join(out_root, f'third_stage_first5_eigenvectors_nmod4_{g}.pdf')}"
            )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())